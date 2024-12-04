import pygame
import random
import os
import time
import math
import json
import glob
import gc
import sys

# Initialize Pygame
pygame.init()

# Screen setup
SCREEN_WIDTH = 800
SCREEN_HEIGHT = 500
screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))
pygame.display.set_caption("RL Game")

# Colors
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
RED = (255, 0, 0)
GREEN = (0, 255, 0)
BLUE = (0, 0, 255)

# Game variables
FPS = 60
GRAVITY = 0.8
JUMP_STRENGTH = -15


class Tile(pygame.sprite.Sprite):
    def __init__(self, x, y, image):
        super().__init__()
        self.image = image
        self.rect = self.image.get_rect()
        self.rect.x = x
        self.rect.y = y

class TileMap:
    def __init__(self):
        self.tile_size = 32
        self.wall_img = pygame.image.load("img/tiles/wall.png").convert_alpha()
        self.ground_img = pygame.image.load("img/tiles/ground.png").convert_alpha()
        self.platform_img = pygame.image.load("img/tiles/platform.png").convert_alpha()
        self.tiles = pygame.sprite.Group()
        self.obstacle_tiles = pygame.sprite.Group()
        self.create_map()

    def create_map(self):
        level = [

            "                        ",
            "                        ",
            "                        ",
            "                              ",
            "                   PPPPP      ",
            "                              ",
            "             PPPPPP           ",
            "                              ",
            "       PPPP                   ",
            "                              ",
            "                              ",
            "                              ",
            "                              ",
            "GGGGGGGGGGGGGGGGGGGGGGGGGGGGGGG"
        ]

        for row, tiles in enumerate(level):
            for col, tile in enumerate(tiles):
                if tile == "W":
                    self.tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.wall_img))
                    self.obstacle_tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.wall_img))
                elif tile == "G":
                    self.tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.ground_img))
                    self.obstacle_tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.ground_img))
                elif tile == "P":
                    self.tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.platform_img))
                    self.obstacle_tiles.add(Tile(col * self.tile_size, row * self.tile_size, self.platform_img))

    def draw(self, surface):
        self.tiles.draw(surface)
        
        
class Character(pygame.sprite.Sprite):
    def __init__(self, x, y):
        super().__init__()
        self.rect = pygame.Rect(x, y, 30, 50)
        self.vel_y = 0
        self.jumping = False
        self.falling = False

    def move(self, dx, tile_map):
        self.rect.x += dx
        for tile in tile_map.obstacle_tiles:
            if self.rect.colliderect(tile.rect):
                if dx > 0:
                    self.rect.right = tile.rect.left
                elif dx < 0:
                    self.rect.left = tile.rect.right

    def jump(self):
        if not self.jumping and not self.falling:
            self.vel_y = JUMP_STRENGTH
            self.jumping = True
            return True
        return False

    def update(self, tile_map):
        self.vel_y += GRAVITY
        self.rect.y += self.vel_y

        for tile in tile_map.obstacle_tiles:
            if self.rect.colliderect(tile.rect):
                if self.vel_y > 0:
                    self.rect.bottom = tile.rect.top
                    self.jumping = False
                    self.falling = False
                    self.vel_y = 0
                elif self.vel_y < 0:
                    self.rect.top = tile.rect.bottom
                    self.vel_y = 0

        if self.vel_y > 0:
            self.falling = True

class Player(Character):
    animation_lists = None

    @classmethod
    def load_animations(cls):
        if cls.animation_lists is None:
            cls.animation_lists = []
            animation_types = ["Idle", "Run", "Jump", "Death", "Attack", "Fall", "Hurt"]
            for animation in animation_types:
                temp_list = []
                num_of_frames = len(os.listdir(f"img/Player/{animation}"))
                for i in range(num_of_frames):
                    img = pygame.image.load(f"img/Player/{animation}/{i}.png").convert_alpha()
                    img = pygame.transform.scale(img, (int(img.get_width() * 2), int(img.get_height() * 2)))
                    temp_list.append(img)
                cls.animation_lists.append(temp_list)

    def __init__(self, x, y):
        super().__init__(x, y)
        if Player.animation_lists is None:
            Player.load_animations()
        self.animation_list = Player.animation_lists
        self.health = 100
        self.max_health = self.health
        self.speed = 6
        self.action = 0  # 0: Idle, 1: Run, 2: Jump, 3: Death, 4: Attack, 5: Fall, 6: Hurt
        self.frame_index = 0
        self.update_time = pygame.time.get_ticks()
        self.attacking = False
        self.attack_cooldown = 0
        self.facing_right = True
        self.alive = True
        self.hit_timer = 0
        self.knockback_speed = 0
        self.image = self.animation_list[self.action][self.frame_index]
        self.rect = self.image.get_rect()
        self.rect.midbottom = (x, y)
        self.shielded = False
        self.shield_blocked_attack = False
        self.attack_range = 50
        self.has_hit_enemy = False

    def update(self, tile_map):
        super().update(tile_map)
        if self.alive:
            
            self.update_animation()
            if self.attack_cooldown > 0:
                self.attack_cooldown -= 1
            
            if self.hit_timer > 0:
                self.hit_timer -= 1
                self.rect.x += self.knockback_speed
                self.knockback_speed *= 0.9  # Decelerate the knockback
            
            if not self.jumping and not self.falling and self.hit_timer == 0:
                if self.action in [2, 5, 6]:  # If was jumping, falling, or hurt
                    self.update_action(0)  # Set to Idle when landing or recovering
            elif self.vel_y > 0 and not self.falling:
                self.falling = True
                self.update_action(5)  # Set to Fall animation
        else:
            # If not alive, only update the death animation
            self.update_death_animation()

    def update_animation(self):
        ANIMATION_COOLDOWN = 100
        self.image = self.animation_list[self.action][self.frame_index]
        if not self.facing_right:
            self.image = pygame.transform.flip(self.image, True, False)
        
        if pygame.time.get_ticks() - self.update_time > ANIMATION_COOLDOWN:
            self.update_time = pygame.time.get_ticks()
            self.frame_index += 1
        if self.frame_index >= len(self.animation_list[self.action]):
            if self.action == 4:  # Attack animation finished
                self.attacking = False
                self.update_action(0)  # Set to Idle after attack
            elif self.action in [2, 5]:  # For jump and fall animations, stay on last frame
                self.frame_index = len(self.animation_list[self.action]) - 1
            elif self.action == 6:  # Hurt animation finished
                self.update_action(0)  # Set to Idle after hurt
            else:
                self.frame_index = 0

    def update_death_animation(self):
        ANIMATION_COOLDOWN = 150  # Slower animation for death
        self.image = self.animation_list[3][self.frame_index]  # 3 is the index for Death animation
        if not self.facing_right:
            self.image = pygame.transform.flip(self.image, True, False)
        
        if pygame.time.get_ticks() - self.update_time > ANIMATION_COOLDOWN:
            self.update_time = pygame.time.get_ticks()
            if self.frame_index < len(self.animation_list[3]) - 1:
                self.frame_index += 1

    def move(self, dx, tile_map):
        if self.alive and not self.attacking and self.hit_timer == 0:  # Only move if alive and not attacking or hurt
            super().move(dx, tile_map)
            if dx != 0:
                self.facing_right = dx > 0
                if not self.jumping and not self.falling:
                    self.update_action(1)  # Set to Run animation only if on the ground
            elif not self.jumping and not self.falling:
                self.update_action(0)  # Set to Idle animation if on the ground and not moving


    def jump(self):
        if self.alive and super().jump():
            self.update_action(2)  # Set to Jump animation
            return True
        return False

    def attack(self):
        if self.alive and self.attack_cooldown == 0 and not self.attacking and not self.jumping and not self.falling and self.hit_timer == 0:
            self.attacking = True
            self.attack_cooldown = 20
            self.update_action(4)  # Set to Attack animation
            self.has_hit_enemy = False
            return True
        return False

    def update_action(self, new_action):
        if self.alive and new_action != self.action:
            self.action = new_action
            self.frame_index = 0
            self.update_time = pygame.time.get_ticks()

    def take_damage(self, amount, knockback_direction):
        if self.alive and not self.shielded:
            self.health -= amount
            if self.health <= 0:
                self.health = 0
                self.alive = False
                self.update_action(3)  # Set to Death animation
                self.frame_index = 0  # Start death animation from the beginning
            else:
                self.hit_timer = 30  # 0.5 seconds at 60 FPS
                self.knockback_speed = knockback_direction * 5  # Adjust for desired knockback strength
                self.update_action(6)  # Set to Hurt animation
                self.attacking = False  # Reset attacking state when hit
                self.attack_cooldown = 0  # Reset attack cooldown when hit
        elif self.shielded:
            self.shield_blocked_attack = True

    def reset(self):
        self.rect.midbottom = (self.initial_x, self.initial_y)
        self.health = self.max_health
        self.alive = True
        self.action = 0  # Set to Idle
        self.frame_index = 0
        self.attacking = False
        self.attack_cooldown = 0
        self.facing_right = True
        self.hit_timer = 0
        self.knockback_speed = 0
        self.jumping = False
        self.falling = False
        self.vel_y = 0
        self.shielded = False
        self.shield_blocked_attack = False
        self.update_time = pygame.time.get_ticks()
        self.image = self.animation_list[self.action][self.frame_index]

    def reset_shield(self):
        self.shielded = False
        self.shield_blocked_attack = False

class Arrow(pygame.sprite.Sprite):
    def __init__(self, x, y, direction):
        super().__init__()
        self.image = pygame.image.load("img/archer/arrow/0.png").convert_alpha()
        self.image = pygame.transform.scale(self.image, (int(self.image.get_width() * 1.5), int(self.image.get_height() * 1.5)))
        self.rect = self.image.get_rect()
        self.rect.center = (x, y)
        self.speed = 6
        self.direction = direction
        self.vel_y = 0
        self.angle = 0
        self.stopped = False

    def update(self):
        if not self.stopped:
            self.vel_y += GRAVITY * 0.05
            self.rect.x += self.speed * self.direction
            self.rect.y += self.vel_y
            
            # Rotate arrow based on trajectory
            self.angle = -math.atan2(self.vel_y, self.speed * self.direction)
            rotated_image = pygame.transform.rotate(pygame.image.load("img/archer/arrow/0.png").convert_alpha(), math.degrees(self.angle))
            self.image = pygame.transform.scale(rotated_image, (int(rotated_image.get_width() * 1.5), int(rotated_image.get_height() * 1.5)))

            # Check if arrow hits the ground
            if self.rect.bottom >= SCREEN_HEIGHT - 60:
                self.rect.bottom = SCREEN_HEIGHT - 60
                self.stopped = True

        # Remove arrow if it goes off-screen horizontally
        if self.rect.right < 0 or self.rect.left > SCREEN_WIDTH:
            self.kill()

class Arrow(pygame.sprite.Sprite):
    def __init__(self, x, y, direction):
        super().__init__()
        self.image = pygame.image.load("img/archer/arrow/0.png").convert_alpha()
        self.image = pygame.transform.scale(self.image, (int(self.image.get_width() * 1.5), int(self.image.get_height() * 1.5)))
        self.rect = self.image.get_rect()
        self.rect.center = (x, y)
        self.speed = 6
        self.direction = direction
        self.vel_y = 0
        self.angle = 0
        self.stopped = False

    def update(self):
        if not self.stopped:
            self.vel_y += GRAVITY * 0.05
            self.rect.x += self.speed * self.direction
            self.rect.y += self.vel_y
            
            # Rotate arrow based on trajectory
            self.angle = -math.atan2(self.vel_y, self.speed * self.direction)
            rotated_image = pygame.transform.rotate(pygame.image.load("img/archer/arrow/0.png").convert_alpha(), math.degrees(self.angle))
            self.image = pygame.transform.scale(rotated_image, (int(rotated_image.get_width() * 1.5), int(rotated_image.get_height() * 1.5)))

            # Check if arrow hits the ground
            if self.rect.bottom >= SCREEN_HEIGHT - 60:
                self.rect.bottom = SCREEN_HEIGHT - 60
                self.stopped = True

        # Remove arrow if it goes off-screen horizontally
        if self.rect.right < 0 or self.rect.left > SCREEN_WIDTH:
            self.kill()

class Enemy(Character):
    animation_lists = None

    @classmethod
    def load_animations(cls):
        if cls.animation_lists is None:
            cls.animation_lists = []
            animation_types = ["Idle", "Run", "Death", "Attack"]
            for animation in animation_types:
                temp_list = []
                num_of_frames = len(os.listdir(f"img/archer/{animation}"))
                for i in range(num_of_frames):
                    img = pygame.image.load(f"img/archer/{animation}/{i}.png").convert_alpha()
                    img = pygame.transform.scale(img, (int(img.get_width() * 1.5), int(img.get_height() * 1.5)))
                    temp_list.append(img)
                cls.animation_lists.append(temp_list)

    def __init__(self, x, y):
        super().__init__(x, y)
        if Enemy.animation_lists is None:
            Enemy.load_animations()
        self.animation_list = Enemy.animation_lists
        self.health = 50
        self.max_health = self.health
        self.previous_health = self.health
        self.speed = 5
        self.direction = 1
        self.action = 0  # 0: Idle, 1: Run, 2: Death, 3: Attack
        self.frame_index = 0
        self.update_time = pygame.time.get_ticks()
        self.alive = True
        self.death_timer = time.time()
        self.vertical_offset = 0
        self.flash_timer = 0
        self.attack_cooldown = 0
        self.arrow_group = pygame.sprite.Group()
        self.attacking = False
        self.attack_frame = 0
        self.invulnerable_timer = 0
        self.invulnerable_duration = 60  # 1 second at 60 FPS
        self.just_attacked = False
        self.death_penalty_applied = False

        self.image = self.animation_list[self.action][self.frame_index]
        self.rect = self.image.get_rect()
        self.rect.x = max(0, min(x, SCREEN_WIDTH - self.rect.width))
        self.rect.bottom = y + self.vertical_offset

        self.sarsa = SARSA(character_type="enemy")
        self.previous_state = None
        self.previous_action = None
        self.episode_steps = 0
        self.total_reward = 0
        self.knockback_velocity = 0
        self.knockback_decay = 0.8

    def get_state(self, player):
        dx = player.rect.x - self.rect.x
        dy = player.rect.y - self.rect.y

        if abs(dx) <= 40:
            x_state = "melee_range"
        elif abs(dx) <= 80:
            x_state = "close"
        elif abs(dx) <= 120:
            x_state = "medium_close"
        elif abs(dx) <= 160:
            x_state = "medium"
        elif abs(dx) <= 200:
            x_state = "medium_far"
        elif abs(dx) <= 250:
            x_state = "far"
        elif abs(dx) <= 300:
            x_state = "very_far"
        elif abs(dx):
            x_state = "extreme_range"

        x_direction = "right" if dx > 0 else "left"

        if abs(dy) <= 50:
            y_state = "same_level"
        elif dy < -50:
            y_state = "above"
        elif 50 < dy:
            y_state = "below"

        enemy_health = "high" if self.health > 35 else ("medium" if self.health > 15 else "low")
        player_health = "high" if player.health > 66 else ("medium" if player.health > 33 else "low")

        facing_player = "facing_player" if (self.direction == 1 and dx > 0) or (
                    self.direction == -1 and dx < 0) else "not_facing_player"

        attack_ready = "attack_ready" if self.attack_cooldown == 0 else "attack_cooldown"


        if self.rect.left <= 100:
            wall_state = "far_to_left_wall"
        elif self.rect.right >= SCREEN_WIDTH - 100:
            wall_state = "far_to_right_wall"
        else:
            wall_state = "no_wall"

        return f"{x_state}_{x_direction}_{y_state}_{enemy_health}_{player_health}_{facing_player}_{attack_ready}_{wall_state}"

    def update_animation(self):
        ANIMATION_COOLDOWN = 100
        max_frames = len(self.animation_list[self.action])
        self.frame_index = min(self.frame_index, max_frames - 1)

        self.image = self.animation_list[self.action][self.frame_index]
        if self.direction == -1:
            self.image = pygame.transform.flip(self.image, True, False)

        if self.flash_timer > 0 and self.flash_timer % 4 < 2:
            self.image = self.image.copy()
            self.image.fill((255, 255, 255, 128), special_flags=pygame.BLEND_RGBA_MULT)
        elif self.invulnerable_timer > 0 and self.invulnerable_timer % 4 < 2:
            self.image = self.image.copy()
            self.image.fill((200, 200, 255, 128), special_flags=pygame.BLEND_RGBA_MULT)

        if self.attacking:
            self.frame_index = self.attack_frame
        elif pygame.time.get_ticks() - self.update_time > ANIMATION_COOLDOWN:
            self.update_time = pygame.time.get_ticks()
            self.frame_index += 1
            if self.frame_index >= max_frames:
                if self.action == 2:
                    self.frame_index = max_frames - 1
                elif self.action == 3:
                    self.frame_index = 0
                    self.update_action(0)
                else:
                    self.frame_index = 0

    def update_action(self, new_action):
        if new_action != self.action:
            self.action = new_action
            self.frame_index = 0
            self.update_time = pygame.time.get_ticks()

    def update(self, player, tile_map):
        super().update(tile_map)
        self.update_animation()
        if self.attack_cooldown > 0:
            self.attack_cooldown -= 1

        if self.invulnerable_timer > 0:
            self.invulnerable_timer -= 1

        if self.flash_timer > 0:
            self.flash_timer -= 1

        if self.alive:
            current_state = self.get_state(player)
            action = self.sarsa.get_action(current_state)

            # Only act if not being knocked back significantly
            if abs(self.knockback_velocity) < 1:
                self.act(action, tile_map)

            # Apply knockback and reduce its effect over time
            if self.knockback_velocity != 0:
                new_x = self.rect.x + int(self.knockback_velocity)
                # Ensure the enemy stays within screen boundaries
                if 0 <= new_x <= SCREEN_WIDTH - self.rect.width:
                    self.rect.x = new_x
                else:
                    # If would go out of bounds, stop at the edge
                    self.rect.x = max(0, min(SCREEN_WIDTH - self.rect.width, new_x))
                    self.knockback_velocity = 0  # Stop knockback if hit boundary

                self.knockback_velocity *= self.knockback_decay
                if abs(self.knockback_velocity) < 0.5:
                    self.knockback_velocity = 0

            self.previous_state = current_state
            self.previous_action = action

            self.episode_steps += 1

        if self.attacking:
            self.attack_frame += 1
            if self.attack_frame >= len(self.animation_list[3]):
                self.attacking = False
                self.attack_frame = 0
                self.shoot_arrow()

        self.arrow_group.update()

    def act(self, action, tile_map):
        if self.alive and self.knockback_velocity == 0:  # Only act if not being knocked back
            if action == 'move_left':
                self.direction = -1
                self.move_ai(tile_map)
            elif action == 'move_right':
                self.direction = 1
                self.move_ai(tile_map)
            elif action == 'shoot':
                self.attack()

    def move_ai(self, tile_map):
        super().update(tile_map)
        if self.alive and not self.attacking:
            new_x = self.rect.x + self.direction * self.speed

            if 0 <= new_x <= SCREEN_WIDTH - self.rect.width:
                self.rect.x = new_x
            else:
                self.direction *= -1

            self.update_action(1)  # Walk animation

    def attack(self):
        if self.attack_cooldown == 0 and self.alive and not self.attacking:
            self.attacking = True
            self.attack_frame = 0
            self.attack_cooldown = 90  # Set cooldown to 90 frames 
            self.update_action(3)
            self.just_attacked = True
            return True
        return False

    def shoot_arrow(self):
        arrow_x = self.rect.centerx + (50 * self.direction)
        arrow_y = self.rect.centery - 10
        new_arrow = Arrow(arrow_x, arrow_y, self.direction)
        self.arrow_group.add(new_arrow)

    def take_damage(self, amount, knockback_direction):
        if self.alive and self.invulnerable_timer == 0:
            self.health -= amount
            self.flash_timer = 30
            self.update_action(0)
            self.invulnerable_timer = self.invulnerable_duration

            # Apply knockback
            self.knockback_velocity = knockback_direction * 10

            if self.health <= 0:
                self.health = 0
                self.alive = False
                self.update_action(2)

    def draw_arrows(self, surface):
        self.arrow_group.draw(surface)

    def check_arrow_hit(self, player):
        hit_player = False
        killed_player = False
        for arrow in self.arrow_group:
            if player.alive and not player.shielded and not arrow.stopped and arrow.rect.colliderect(player.rect):
                knockback_direction = 1 if arrow.direction > 0 else -1
                player.take_damage(10, knockback_direction)
                arrow.kill()
                hit_player = True
                if not player.alive:
                    killed_player = True
                break
        return hit_player, killed_player

    def reset(self):
        self.health = self.max_health
        self.previous_health = self.health
        self.rect.x = 500
        self.rect.y = SCREEN_HEIGHT - 200
        self.alive = True
        self.action = 0
        self.frame_index = 0
        self.arrow_group.empty()
        self.attacking = False
        self.attack_frame = 0
        self.flash_timer = 0
        self.attack_cooldown = 0
        self.episode_steps = 0
        self.total_reward = 0
        self.previous_state = None
        self.previous_action = None
        self.invulnerable_timer = 0
        self.just_attacked = False
        self.death_penalty_applied = False
        self.knockback_velocity = 0

    def end_episode(self):
        self.sarsa.end_episode()

        self.episode_steps = 0
        self.total_reward = 0

        
        
class AIPlayer(Player):
    def __init__(self, x, y):
        super().__init__(x, y)
        self.decision_cooldown = 0
        self.attack_idle_time = 0
        self.has_hit_enemy = False  # New attribute to track if the current attack has hit the enemy

    def make_decision(self, enemy):
        if self.attack_idle_time > 0:
            self.attack_idle_time -= 1
            return

        if self.decision_cooldown > 0:
            self.decision_cooldown -= 1
            return

        dx = enemy.rect.centerx - self.rect.centerx

        # Move towards the enemy
        if abs(dx) > 45:  # Approach until we're in attack range
            if dx > 0:
                self.move(self.speed)
            else:
                self.move(-self.speed)
        else:
            # If in range, attack
            if random.random() < 0.8:  # 80% chance to attack when in range
                if self.attack():  # Only set idle time if attack was successful
                    self.attack_idle_time = 10  # 0.5 seconds at 60 FPS
            else:
                # Sometimes move slightly away to avoid getting hit
                self.move(-self.speed if dx > 0 else self.speed)

        self.decision_cooldown = 3  # Wait 3 frames before next decision

    def update(self, enemy, tile_map):
        super().update(tile_map)
        self.make_decision(enemy)

        # Check for attack hitting enemy
        if self.attacking and not self.has_hit_enemy:
            if (abs(self.rect.centerx - enemy.rect.centerx) < 50 and
                abs(self.rect.centery - enemy.rect.centery) < 50):
                knockback_direction = 1 if self.facing_right else -1
                enemy.take_damage(5, knockback_direction)
                self.has_hit_enemy = True

        # Reset has_hit_enemy when attack animation ends
        if not self.attacking:
            self.has_hit_enemy = False

    def move(self, dx):
        # Override the move method to use the full player speed
        if self.alive and not self.attacking and self.hit_timer == 0:
            self.rect.x += dx
            self.rect.x = max(0, min(self.rect.x, SCREEN_WIDTH - self.rect.width))  # Keep player within screen bounds
            if dx != 0:
                self.facing_right = dx > 0
                if not self.jumping and not self.falling:
                    self.update_action(1)  # Set to Run animation
            elif not self.jumping and not self.falling:
                self.update_action(0)  # Set to Idle animation

    def attack(self):
        if super().attack():  # Call the parent class's attack method
            self.attack_idle_time = 30  # Set idle time after successful attack
            self.has_hit_enemy = False  # Reset has_hit_enemy at the start of a new attack
            return True
        return False


    def reset(self):
        self.rect.x = random.randint(50, 750)
        self.rect.bottom = SCREEN_HEIGHT - 50
        self.health = self.max_health
        self.alive = True
        self.action = 0
        self.frame_index = 0
        self.attacking = False
        self.jumping = False
        self.falling = False
        self.vel_y = 0
        self.facing_right = True
        self.decision_cooldown = 0
        self.attack_idle_time = 0
        self.has_hit_enemy = False



        
        
        
class SARSA:
    def __init__(self, character_type):
        self.character_type = character_type
        self.epsilon = 0
        self.epsilon_decay = 0.999997
        self.epsilon_min = 0.0
        self.alpha = 0.1
        self.alpha_decay = 0.9999
        self.alpha_min = 0.01
        self.gamma = 0.9

        if character_type == "knight":
            self.actions = ['move_left', 'move_right', 'attack', 'block', 'maintain_block', 'idle']
            self.q_table_folder = 'knight_q_tables'
        elif character_type == "enemy":
            self.actions = ['move_left', 'move_right', 'shoot', 'idle']
            self.q_table_folder = 'q_tables'
        elif character_type == "bird":
            self.actions = ['move_up', 'move_down', 'move_left', 'move_right', 'move_up_left', 'move_up_right', 'move_down_left', 'move_down_right', 'activate_shield', 'idle']
            self.q_table_folder = 'bird_q_tables'
        elif character_type == "rogue":
            self.actions = ['move_left', 'move_right', 'far_attack', 'close_attack', 'idle']
            self.q_table_folder = 'rogue_q_tables'
        else:
            raise ValueError(f"Unknown character type: {character_type}")

        self.q_table = self.load_q_table()
        self.episode_count = self.get_latest_episode_count()

    def get_latest_episode_count(self):
        q_table_files = glob.glob(f'{self.q_table_folder}/*.json')
        if not q_table_files:
            return 0
        latest_file = max(q_table_files, key=os.path.getctime)
        return int(latest_file.split('_')[-1].split('.')[0]) + 1

    def load_q_table(self):
        q_table_files = glob.glob(f'{self.q_table_folder}/*.json')
        if not q_table_files:
            return {}
        
        # Extract episode numbers from filenames and find the highest
        episode_numbers = []
        for file in q_table_files:
            try:
                episode_num = int(file.split('_')[-1].split('.')[0])
                episode_numbers.append((episode_num, file))
            except ValueError:
                continue
        
        if not episode_numbers:
            return {}
        
        # Get the file with the highest episode number
        latest_file = max(episode_numbers, key=lambda x: x[0])[1]
        print(f"Loading Q-table from: {latest_file}")  # Added for debugging
        
        with open(latest_file, 'r') as f:
            return json.load(f)

    def save_q_table(self):
        if not os.path.exists(self.q_table_folder):
            os.makedirs(self.q_table_folder)
        filename = f'{self.q_table_folder}/q_table_episode_{self.episode_count}.json'
        with open(filename, 'w') as f:
            json.dump(self.q_table, f, indent=2)
        print(f"Q-table saved as {filename}")

    def get_action(self, state):
        if state not in self.q_table:
            self.q_table[state] = {a: 0 for a in self.actions}
        
        if random.random() < self.epsilon:
            return random.choice(self.actions)
        else:
            return max(self.q_table[state], key=self.q_table[state].get)

    def update_q_table(self, state, action, reward, next_state, next_action):
        if state not in self.q_table:
            self.q_table[state] = {a: 0 for a in self.actions}
        if next_state not in self.q_table:
            self.q_table[next_state] = {a: 0 for a in self.actions}

        current_q = self.q_table[state][action]
        next_q = self.q_table[next_state][next_action]
        new_q = current_q + self.alpha * (reward + self.gamma * next_q - current_q)
        self.q_table[state][action] = new_q

    def get_best_action(self, state):
        if state not in self.q_table:
            self.q_table[state] = {a: 0 for a in self.actions}
        return max(self.q_table[state], key=self.q_table[state].get)
    
    def end_episode(self):
        self.episode_count += 1
        self.epsilon = max(self.epsilon * self.epsilon_decay, self.epsilon_min)
        #self.alpha = max(self.alpha * self.alpha_decay, self.alpha_min)
        
class SimplePlayer(Player):
    def __init__(self, x, y):
        super().__init__(x, y)
        self.initial_x = x
        self.initial_y = y
        self.direction = 1
        self.speed = 3

    def update(self, tile_map):
        super().update(tile_map)
        self.move(self.direction * self.speed, tile_map)
        
        # Change direction when reaching screen borders
        if self.rect.left <= 0:
            self.direction = 1
        elif self.rect.right >= SCREEN_WIDTH:
            self.direction = -1

    def reset(self):
        super().reset()
        self.direction = 1
        self.rect.x = self.initial_x
        self.rect.y = self.initial_y

class Bird(pygame.sprite.Sprite):
    def __init__(self, x, y):
        super().__init__()
        self.load_animations()
        self.rect = self.image.get_rect()
        self.rect.center = (x, y)
        self.speed = 3
        self.heal_cooldown = 0
        self.heal_cooldown_max = 300
        self.state = "idle"
        self.frame_index = 0
        self.update_time = pygame.time.get_ticks()
        self.facing_right = True
        
        self.sarsa = SARSA(character_type="bird")
        self.previous_state = None
        self.previous_action = None
        self.total_reward = 0
        self.previous_distance = float('inf')
        
        # Shield-related attributes
        self.shield_cooldown = 0
        self.shield_cooldown_max = 390  # 6.5 seconds at 60 FPS
        self.shield_duration = 90  # 1.5 seconds at 60 FPS
        self.shield_active = False
        self.shield_loading = False
        self.shield_frame = 0
        self.shield_animations = {
            "loading": self.load_animation("shield/loading", scale=1),
            "working": self.load_animation("shield/working", scale=1)
        }
        self.shield_reward_given = False
        self.shield_start_time = 0
        self.unnecessary_shield_use = False
        
    def load_animations(self):
        self.animations = {
            "idle": self.load_animation("bird", scale=0.05),
        }
        self.image = self.animations["idle"][0]
        
    def load_animation(self, folder, scale=1):
        animation = []
        for i in range(len(os.listdir(f"img/{folder}"))):
            img = pygame.image.load(f"img/{folder}/{i}.png").convert_alpha()
            img = pygame.transform.scale(img, (int(img.get_width() * scale), int(img.get_height() * scale)))
            animation.append(img)
        return animation
        
    def update(self, player, enemy, knight):
        self.heal_cooldown = max(0, self.heal_cooldown - 1)
        self.shield_cooldown = max(0, self.shield_cooldown - 1)
        
        current_state = self.get_state(player)
        action = self.sarsa.get_action(current_state)
        self.perform_action(action, player)
        
        reward = self.get_reward(player, knight=knight, enemy=enemy)
        self.total_reward += reward
        next_state = self.get_state(player)
        next_action = self.sarsa.get_action(next_state)

        if self.previous_state is not None and self.previous_action is not None:
            self.sarsa.update_q_table(self.previous_state, self.previous_action, reward, current_state, action)

        self.previous_state = current_state
        self.previous_action = action
        
        self.update_animation()
        self.update_shield(player)
        
    def get_state(self, player, knight=None, enemy=None):
        dx = abs(self.rect.centerx - player.rect.centerx)
        dy = player.rect.top - self.rect.bottom  # Positive when bird is above player
        
        if dx <= 100 and dy <= 100:
            proximity = "close"
        elif dx <= 150 and dy <= 150:
            proximity = "far"
        else:
            proximity = "very_far"
        
        x_direction = "right" if player.rect.centerx > self.rect.centerx else "left"
        y_direction = "above" if player.rect.centery < self.rect.centery else "below"
        
        shield_state = "shield_active" if self.shield_active else "shield_inactive"
        shield_cooldown = "shield_ready" if self.shield_cooldown == 0 else "shield_cooldown"

        # Calculate distances and actions for knight
        if knight:
            player_to_knight_distance = abs(player.rect.centerx - knight.rect.centerx)
            if player_to_knight_distance <= 60:
                pk_distance = "very_close"
            elif player_to_knight_distance <= 100:
                pk_distance = "close"
            elif player_to_knight_distance <= 200:
                pk_distance = "medium"
            else:
                pk_distance = "far"
            knight_action = ["idle", "attack", "walk", "death", "block"][knight.action]
        else:
            pk_distance = "far"
            knight_action = "idle"

        # Calculate distances and actions for enemy
        if enemy:
            player_to_enemy_distance = abs(player.rect.centerx - enemy.rect.centerx)
            if player_to_enemy_distance <= 50:
                pe_distance = "very_close"
            elif player_to_enemy_distance <= 100:
                pe_distance = "close"
            elif player_to_enemy_distance <= 200:
                pe_distance = "medium"
            else:
                pe_distance = "far"
            enemy_action = ["idle", "run", "death", "attack"][enemy.action]
        else:
            pe_distance = "far"
            enemy_action = "idle"

        return f"{proximity}_{x_direction}_{y_direction}_{shield_state}_{shield_cooldown}_{pk_distance}_{pe_distance}_{knight_action}_{enemy_action}"
        
    def perform_action(self, action, player):
        dx, dy = 0, 0
        if action == 'move_up': dy = -self.speed
        elif action == 'move_down': dy = self.speed
        elif action == 'move_left': 
            dx = -self.speed
            self.facing_right = False
        elif action == 'move_right': 
            dx = self.speed
            self.facing_right = True
        elif action == 'move_up_left': 
            dx, dy = -self.speed, -self.speed
            self.facing_right = False
        elif action == 'move_up_right': 
            dx, dy = self.speed, -self.speed
            self.facing_right = True
        elif action == 'move_down_left': 
            dx, dy = -self.speed, self.speed
            self.facing_right = False
        elif action == 'move_down_right': 
            dx, dy = self.speed, self.speed
            self.facing_right = True
        elif action == 'activate_shield':
            self.activate_shield(player)
        
        self.rect.x += dx
        self.rect.y += dy
        
        # Keep bird within screen bounds
        self.rect.clamp_ip(pygame.Rect(0, 0, SCREEN_WIDTH, SCREEN_HEIGHT))
        
    def get_reward(self, player, knight=None, enemy=None):
        reward = 0
        
        # Calculate horizontal and vertical distances to the player
        dx = self.rect.x - player.rect.x
        dy = player.rect.y - self.rect.y  # Positive when bird is above player

        # Reward for horizontal proximity
        if dx <= 150:
            reward += 0.5  # Good reward for being close horizontally
        else:
            reward -= 0.1  # Penalty for being too far horizontally

        # Reward for vertical position
        if 130 <= dy <= 170:
            reward += 1  #  reward for being in the ideal vertical range
        else:
            reward -= 1  # Small penalty for being too high

        # Reward for successful shielding
        if player.shielded and not self.shield_reward_given:

            if enemy:
                for arrow in enemy.arrow_group:
                    if not arrow.stopped and arrow.rect.colliderect(player.rect):
                        reward += 50  # Big reward for blocking an attack
                        self.shield_reward_given = True
                        self.unnecessary_shield_use = False
                        break
            # elif knight:
            #     if knight.attacking and abs(knight.rect.centerx - player.rect.centerx) < knight.attack_range:
            #         reward += 50  # Big reward for blocking an attack
            #         self.shield_reward_given = True
            #         self.unnecessary_shield_use = False
            else:
                self.unnecessary_shield_use = True

        # Penalty for unnecessary shield use
        if self.unnecessary_shield_use and not player.shielded and time.time() - self.shield_start_time >= self.shield_duration / 60:
            reward -= 5  # Small penalty for unnecessary shield use
            self.unnecessary_shield_use = False

        return reward

    def update_animation(self):
        ANIMATION_COOLDOWN = 100
        self.image = self.animations[self.state][self.frame_index]
        
        if not self.facing_right:
            self.image = pygame.transform.flip(self.image, True, False)
        
        if pygame.time.get_ticks() - self.update_time > ANIMATION_COOLDOWN:
            self.update_time = pygame.time.get_ticks()
            self.frame_index += 1
        if self.frame_index >= len(self.animations[self.state]):
            self.frame_index = 0

    def activate_shield(self, player):
        if self.shield_cooldown == 0 and not self.shield_active and not self.shield_loading:
            self.shield_loading = True
            self.shield_frame = 0
            player.shielded = True
            self.shield_reward_given = False
            self.shield_start_time = time.time()
            self.unnecessary_shield_use = False

    def update_shield(self, player):
        if self.shield_loading:
            self.shield_frame += 1
            if self.shield_frame >= len(self.shield_animations["loading"]) * 2:
                self.shield_loading = False
                self.shield_active = True
                self.shield_frame = 0
        elif self.shield_active:
            self.shield_frame += 1
            if self.shield_frame >= self.shield_duration:
                self.shield_active = False
                player.shielded = False
                self.shield_cooldown = self.shield_cooldown_max
                if self.unnecessary_shield_use:
                    self.total_reward -= 5  # Apply penalty for unnecessary shield use
                self.unnecessary_shield_use = False

    def draw_shield(self, screen, player):
        if self.shield_loading:
            shield_image = self.shield_animations["loading"][self.shield_frame // 2 % len(self.shield_animations["loading"])]
            screen.blit(shield_image, (self.rect.centerx - shield_image.get_width() // 2, self.rect.top - shield_image.get_height()))
        elif self.shield_active:
            shield_image = self.shield_animations["working"][self.shield_frame // 6 % len(self.shield_animations["working"])]
            shield_image.set_alpha(128)  # Make it semi-transparent
            screen.blit(shield_image, (player.rect.centerx - shield_image.get_width() // 2, player.rect.centery - shield_image.get_height() // 2))

    def reset(self):
        self.rect.center = (400, SCREEN_HEIGHT - 100)
        self.heal_cooldown = 0
        self.state = "idle"
        self.frame_index = 0
        self.previous_state = None
        self.previous_action = None
        self.total_reward = 0
        self.facing_right = True
        self.previous_distance = float('inf')
        self.shield_cooldown = 0
        self.shield_active = False
        self.shield_loading = False
        self.shield_frame = 0
        self.shield_reward_given = False
        self.unnecessary_shield_use = False

    def end_episode(self):
        self.sarsa.end_episode()
        #print(f"Episode {self.sarsa.episode_count} ended. Total Reward: {self.total_reward}")
        print(f"Epsilon: {self.sarsa.epsilon:.6f}, Alpha: {self.sarsa.alpha:.6f}")
        self.total_reward = 0


class Knight(Character):
    animation_lists = None

    @classmethod
    def load_animations(cls):
        if cls.animation_lists is None:
            cls.animation_lists = []
            animation_types = ["Idle", "Attack", "Walk", "Death", "Block"]
            for animation in animation_types:
                temp_list = []
                num_of_frames = len(os.listdir(f"C:\\Users\\mikol\\Desktop\\Reinforcement Learning\\img\\knight\\{animation}"))
                for i in range(num_of_frames):
                    img = pygame.image.load(f"C:\\Users\\mikol\\Desktop\\Reinforcement Learning\\img\\knight\\{animation}\\{i}.png").convert_alpha()
                    img = pygame.transform.scale(img, (int(img.get_width() * 2), int(img.get_height() * 2)))
                    temp_list.append(img)
                cls.animation_lists.append(temp_list)

    def __init__(self, x, y):
        super().__init__(x, y)
        if Knight.animation_lists is None:
            Knight.load_animations()
        self.animation_list = Knight.animation_lists
        self.health = 100
        self.max_health = self.health
        self.previous_health = self.health
        self.speed = 3
        self.direction = 1
        self.action = 0  # 0: Idle, 1: Attack, 2: Walk, 3: Death, 4: Block
        self.frame_index = 0
        self.update_time = pygame.time.get_ticks()
        self.alive = True
        self.death_timer = time.time()
        self.vertical_offset = 0
        self.flash_timer = 0
        self.attack_cooldown = 0
        self.attacking = False
        self.blocking = False
        self.attack_frame = 0
        self.invulnerable_timer = 0
        self.invulnerable_duration = 60  # 1 second at 60 FPS
        self.just_attacked = False
        self.death_penalty_applied = False
        self.shield_used = False
        self.attack_landed = False
        
        self.image = self.animation_list[self.action][self.frame_index]
        self.rect = self.image.get_rect()
        self.rect.x = max(0, min(x, SCREEN_WIDTH - self.rect.width))
        self.rect.bottom = y + self.vertical_offset
        
        self.sarsa = SARSA(character_type="knight")
        self.previous_state = None
        self.previous_action = None
        self.episode_steps = 0
        self.total_reward = 0
        self.knockback_velocity = 0
        self.knockback_decay = 0.7
        self.shield_cooldown = 0
        self.shield_cooldown_max = 60  # 1 second at 60 FPS
        
        self.attack_range = 60  # Updated attack range
        self.block_duration = 0
        self.max_block_duration = 120  # 2 seconds at 60 FPS
        self.block_release_cooldown = 60  # 1 second cooldown after releasing block
        self.player = None
        
        self.hit_player = False
        self.killed_player = False

    def get_state(self, player):
        dx = player.rect.x - self.rect.x
        dy = player.rect.y - self.rect.y
        
        if abs(dx) <= self.attack_range:
            x_state = "melee_range"
        elif abs(dx) <= 100:
            x_state = "close"
        elif abs(dx) <= 200:
            x_state = "medium"
        else:
            x_state = "far"
        
        x_direction = "right" if dx > 0 else "left"
        
        if abs(dy) <= 50:
            y_state = "same_level"
        elif dy < -50:
            y_state = "above"
        else:
            y_state = "below"
        
        knight_health = "high" if self.health > 66 else ("medium" if self.health > 20 else "low")
        player_health = "high" if player.health > 66 else ("medium" if player.health > 20 else "low")
        
        current_action = ["idle", "attack", "walk", "death", "block"][self.action]
        
        facing_player = "facing_player" if self.is_facing_player() else "not_facing_player"
        
        attack_ready = "attack_ready" if self.attack_cooldown == 0 else "attack_cooldown"

        player_attacking = "player_attacking" if player.attacking else "player_not_attacking"

        if self.rect.left <= 50:
            wall_state = "close_to_left_wall"
        elif self.rect.right >= SCREEN_WIDTH - 50:
            wall_state = "close_to_right_wall"
        else:
            wall_state = "no_wall"
            
        shield_ready = "shield_ready" if self.shield_cooldown == 0 else "shield_cooldown"
        
        block_state = f"blocking_{self.block_duration // 30}" if self.blocking else "not_blocking"
        
        return f"{x_state}_{x_direction}_{y_state}_{knight_health}_{player_health}_{current_action}_{facing_player}_{attack_ready}_{player_attacking}_{wall_state}_{shield_ready}_{block_state}"

    def is_facing_player(self):
        if self.player:
            return (self.direction == 1 and self.player.rect.centerx > self.rect.centerx) or \
                   (self.direction == -1 and self.player.rect.centerx < self.rect.centerx)
        return False

    def act(self, action, player, tile_map):
        self.player = player
        if self.attacking:
            return  # Don't perform any new actions while attacking
        
        if action == 'move_left':
            self.direction = -1
            self.move_ai(tile_map)
        elif action == 'move_right':
            self.direction = 1
            self.move_ai(tile_map)
        elif action == 'attack':
            self.attack()
        elif action == 'block':
            if not self.blocking:
                self.block()
            elif self.block_duration >= self.max_block_duration:
                self.release_block()
        elif action == 'maintain_block':
            if self.blocking and self.block_duration < self.max_block_duration and self.is_facing_player():
                self.block_duration += 1
            else:
                self.release_block()
        elif action == 'idle':
            if self.blocking:
                self.release_block()
            else:
                self.update_action(0)
        
        self.previous_action = action

    def move_ai(self, tile_map):
        if self.alive and not self.attacking and not self.blocking:
            dx = self.direction * self.speed
            self.move(dx, tile_map)
            self.update_action(2)  # Walk animation

    def attack(self):
        if self.attack_cooldown == 0 and self.alive and not self.attacking and not self.blocking:
            self.attacking = True
            self.attack_frame = 0
            self.attack_cooldown = 60
            self.update_action(1)
            self.attack_landed = False
            self.just_attacked = True
            return True
        return False

    def block(self):
        if (self.alive and not self.attacking and not self.blocking and 
            self.shield_cooldown == 0 and self.is_facing_player()):
            self.blocking = True
            self.update_action(4)  # Set to Block animation
            self.frame_index = len(self.animation_list[4]) - 1  # Set to last frame of Block animation
            self.shield_used = False
            self.block_duration = 0
            return True
        return False

    def release_block(self):
        if self.blocking:
            self.blocking = False
            self.shield_cooldown = self.block_release_cooldown
            self.update_action(0)

    def take_damage(self, amount, knockback_direction):
        if self.alive and self.invulnerable_timer == 0:
            if not (self.blocking and self.is_facing_player()):
                self.health -= amount
                self.flash_timer = 30
                self.update_action(0)
                self.invulnerable_timer = self.invulnerable_duration
                
                # Apply stronger knockback
                self.knockback_velocity = knockback_direction * 15
            else:
                # Even when blocking, apply some knockback
                self.knockback_velocity = knockback_direction * 5
            
            if self.health <= 0:
                self.health = 0
                self.alive = False
                self.update_action(3)  # Death animation
                if not self.death_penalty_applied:
                    self.total_reward -= 100  # Apply death penalty
                    self.death_penalty_applied = True

    def check_melee_hit(self, player):
        if self.attacking and not self.attack_landed and self.is_facing_player() and \
           abs(self.rect.centerx - player.rect.centerx) < self.attack_range and \
           abs(self.rect.centery - player.rect.centery) < 50:
            knockback_direction = 1 if self.direction > 0 else -1
            player.take_damage(10, knockback_direction)
            self.attack_landed = True
            self.hit_player = True
            if not player.alive:
                self.killed_player = True
            return True
        return False

    def get_reward(self):
        reward = 0
        
        if self.hit_player:
            reward += 30
            if self.killed_player:
                reward += 50  # Additional reward for killing the player
        
        if self.health < self.previous_health:
            reward -= 20
        
        if self.blocking and self.shield_used and self.is_facing_player():
            reward += 30
            
        if self.health == 0:
            if not self.death_penalty_applied:
                reward -= 50  # Apply death penalty
                self.death_penalty_applied = True        
        
        self.previous_health = self.health
        return reward

    def update(self, player, tile_map):
        super().update(tile_map)
        self.update_animation()
        self.hit_player = False
        self.killed_player = False
        self.update_animation()
        if self.attack_cooldown > 0:
            self.attack_cooldown -= 1
        
        if self.invulnerable_timer > 0:
            self.invulnerable_timer -= 1
        
        if self.flash_timer > 0:
            self.flash_timer -= 1
        
        if self.shield_cooldown > 0:
            self.shield_cooldown -= 1
        
        if self.alive:
            current_state = self.get_state(player)
            action = self.sarsa.get_action(current_state)
            self.act(action, player, tile_map)

            self.check_melee_hit(player)
            

            self.previous_state = current_state
            self.previous_action = action

            self.episode_steps += 1

        if self.attacking:
            self.attack_frame += 0.5  # Changed from 1 to 0.5 to slow down the animation
            if self.attack_frame >= len(self.animation_list[1]):
                self.attacking = False
                self.attack_frame = 0

        if self.blocking:
            if self.block_duration < self.max_block_duration and self.is_facing_player():
                self.block_duration += 1
            else:
                self.release_block()
                
        # Apply knockback
        if self.knockback_velocity != 0:
            self.rect.x += int(self.knockback_velocity)
            self.knockback_velocity *= self.knockback_decay
            if abs(self.knockback_velocity) < 0.1:
                self.knockback_velocity = 0
                
        self.rect.x = max(0, min(self.rect.x, SCREEN_WIDTH - self.rect.width))

    def update_animation(self):
        ANIMATION_COOLDOWN = 100
        max_frames = len(self.animation_list[self.action])
        
        if self.blocking:
            self.frame_index = max_frames - 1  # Stay on the last frame of block animation
        else:
            if self.attacking:
                self.frame_index = int(min(self.attack_frame, max_frames - 1))  # Cast to int
            elif pygame.time.get_ticks() - self.update_time > ANIMATION_COOLDOWN:
                self.update_time = pygame.time.get_ticks()
                self.frame_index += 1
                if self.frame_index >= max_frames:
                    if self.action == 3:  # Death animation
                        self.frame_index = max_frames - 1
                    elif self.action in [1, 4]:  # Attack or Block animations
                        self.frame_index = 0
                        self.update_action(0)  # Return to Idle
                    else:
                        self.frame_index = 0

        # Ensure frame_index is always within bounds and an integer
        self.frame_index = int(min(self.frame_index, max_frames - 1))
        
        self.image = self.animation_list[self.action][self.frame_index]
        if self.direction == -1:
            self.image = pygame.transform.flip(self.image, True, False)
        
        if self.flash_timer > 0 and self.flash_timer % 4 < 2:
            self.image = self.image.copy()
            self.image.fill((255, 255, 255, 128), special_flags=pygame.BLEND_RGBA_MULT)

    def update_action(self, new_action):
        if new_action != self.action:
            self.action = new_action
            self.frame_index = 0
            self.update_time = pygame.time.get_ticks()

    def reset(self):
        self.health = self.max_health
        self.previous_health = self.health
        self.rect.x = 500
        self.rect.y = SCREEN_HEIGHT - 200
        self.alive = True
        self.action = 0
        self.frame_index = 0
        self.attacking = False
        self.blocking = False
        self.attack_frame = 0
        self.flash_timer = 0
        self.attack_cooldown = 0
        self.episode_steps = 0
        self.total_reward = 0
        self.previous_state = None
        self.previous_action = None
        self.invulnerable_timer = 0
        self.death_penalty_applied = False 
        self.shield_used = False
        self.attack_landed = False
        self.shield_cooldown = 0
        self.just_attacked = False
        self.block_duration = 0
        self.player = None
        self.hit_player = False
        self.killed_player = False

    def end_episode(self):
        self.sarsa.end_episode()
        self.episode_steps = 0
        self.total_reward = 0

def train_knight_fast():
    tile_map = TileMap()
    knight = Knight(500, SCREEN_HEIGHT - 72)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)

    num_episodes = 500000  
    frames_per_episode = 30 * 60  # 60 seconds at 60 FPS
    FULL_RESET_INTERVAL = 50000

    start_time = time.time()
    last_epsilon = knight.sarsa.epsilon

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            last_epsilon = knight.sarsa.epsilon
            del knight
            gc.collect()
            knight = Knight(500, SCREEN_HEIGHT - 72)
            knight.sarsa.epsilon = last_epsilon
            print(f"Performed full reset at episode {episode}, continuing with epsilon {last_epsilon:.6f}")
        
        knight.reset()
        player.reset()
        frame_count = 0
        episode_reward = 0
        
        while frame_count < frames_per_episode:
            current_state = knight.get_state(player)
            action = knight.sarsa.get_action(current_state)
            
            previous_health = knight.health
            knight.act(action, player, tile_map)
            player.update(knight, tile_map)
            knight.update(player, tile_map)

            # Calculate reward
            dx = player.rect.x - knight.rect.x

            # Calculate reward
            reward = 0
            if knight.just_attacked:
                reward -= 10
                knight.just_attacked = False

            if abs(dx) > 100:
                reward += 0.01
            if knight.hit_player:
                reward += 40
                if knight.killed_player:
                    reward += 100
            if knight.health < previous_health:
                reward -= 40
            if knight.blocking and knight.shield_used and knight.is_facing_player():
                reward += 30
            if knight.health == 0 and not knight.death_penalty_applied:
                reward -= 100
                knight.death_penalty_applied = True
            
            episode_reward += reward

            next_state = knight.get_state(player)
            next_action = knight.sarsa.get_action(next_state)

            # Update Q-table
            knight.sarsa.update_q_table(current_state, action, reward, next_state, next_action)

            frame_count += 1

            if not player.alive or not knight.alive:
                break

        knight.sarsa.end_episode()
        
        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {knight.sarsa.epsilon:.6f}", flush=True)
        
        if (episode + 1) % 100 == 0:
            knight.sarsa.save_q_table()
            gc.collect()

    print("Training complete")
    print(f"Final Epsilon: {knight.sarsa.epsilon:.6f}")
    knight.sarsa.save_q_table()

    total_time = (time.time() - start_time) / 60
    print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()

def visualize_training():
    tile_map = TileMap()
    knight = Knight(70, SCREEN_HEIGHT - 72)
    player = AIPlayer(400, SCREEN_HEIGHT - 50)
    all_sprites = pygame.sprite.Group(player, knight)

    num_episodes = 1000
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS
    FULL_RESET_INTERVAL = 100

    start_time = time.time()
    last_epsilon = knight.sarsa.epsilon

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            last_epsilon = knight.sarsa.epsilon
            del knight
            gc.collect()
            knight = Knight(500, SCREEN_HEIGHT - 72)
            knight.sarsa.epsilon = last_epsilon
            print(f"Performed full reset at episode {episode}, continuing with epsilon {last_epsilon:.6f}")
        
        knight.reset()
        player.reset()
        frame_count = 0
        episode_reward = 0
        
        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            current_state = knight.get_state(player)
            action = knight.sarsa.get_action(current_state)
            
            previous_health = knight.health
            knight.act(action, player, tile_map)
            player.update(knight, tile_map)
            knight.update(player, tile_map)
            dx = player.rect.x - knight.rect.x
            # Calculate reward
            # Calculate reward
            reward = 0
            if knight.just_attacked:
                reward -= 10
                knight.just_attacked = False

            if  abs(dx) > 100:
                reward -= 0.01
            if knight.hit_player:
                reward += 40
                if knight.killed_player:
                    reward += 100
            if knight.health < previous_health:
                reward -= 40
            if knight.blocking and knight.shield_used and knight.is_facing_player():
                reward += 30
            if knight.health == 0 and not knight.death_penalty_applied:
                reward -= 100
                knight.death_penalty_applied = True
            
            episode_reward += reward

            next_state = knight.get_state(player)
            next_action = knight.sarsa.get_action(next_state)

            # Update Q-table
            knight.sarsa.update_q_table(current_state, action, reward, next_state, next_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)
            all_sprites.draw(screen)

            # Draw health bars
            pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 10, player.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (player.rect.x, player.rect.y - 10, player.rect.width * (player.health / player.max_health), 5))
            pygame.draw.rect(screen, RED, (knight.rect.x, knight.rect.y - 10, knight.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (knight.rect.x, knight.rect.y - 10, knight.rect.width * (knight.health / knight.max_health), 5))

            # Display episode number, current reward, and epsilon
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode+1}", True, BLACK)
            reward_text = font.render(f"Reward: {episode_reward:.2f}", True, BLACK)
            epsilon_text = font.render(f"Epsilon: {knight.sarsa.epsilon:.4f}", True, BLACK)
            action_text = font.render(f"Knight Action: {action}", True, BLACK)
            screen.blit(episode_text, (10, 10))
            screen.blit(reward_text, (10, 40))
            screen.blit(epsilon_text, (10, 70))
            screen.blit(action_text, (10, 100))

            pygame.display.flip()
            clock.tick(60)  

            frame_count += 1

            if not player.alive or not knight.alive:
                break

        knight.sarsa.end_episode()
        
        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {knight.sarsa.epsilon:.6f}", flush=True)
        
        if (episode + 1) % 100 == 0:
            knight.sarsa.save_q_table()
            gc.collect()

    print("Visualization complete")
    print(f"Final Epsilon: {knight.sarsa.epsilon:.6f}")
    knight.sarsa.save_q_table()

    total_time = (time.time() - start_time) / 60
    print(f"\nTotal visualization time: {total_time:.2f} minutes")

    gc.collect()
    pygame.quit()
    
def train_bird_with_knight_fast():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    knight = Knight(500, SCREEN_HEIGHT - 72)
    knight.sarsa.epsilon_min = 0
    knight.sarsa.epsilon = 0
    knight.sarsa.q_table = knight.sarsa.load_q_table()
    bird.sarsa.q_table = bird.sarsa.load_q_table()

    num_episodes = 1000000
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS
    FULL_RESET_INTERVAL = 1000

    start_time = time.time()

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            gc.collect()
            print(f"Performed full reset at episode {episode}, continuing with epsilon {bird.sarsa.epsilon:.6f}")
        
        bird.reset()
        player.reset()
        player.reset_shield()
        knight.reset()
        
        frame_count = 0
        episode_reward = 0
        
        while frame_count < frames_per_episode:
            # Knight's turn (using best action, not training)
            knight_state = knight.get_state(player)
            knight_action = knight.sarsa.get_best_action(knight_state)
            knight.act(knight_action, player, tile_map)
            
            # Player's turn
            player.make_decision(knight)
            
            # Bird's turn
            bird_state = bird.get_state(player, knight)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            
            # Update all entities
            player.update(knight,tile_map)
            knight.update(player, tile_map)
            bird.update(player, knight)
            
            # Calculate reward
            reward = bird.get_reward(player, knight)
            episode_reward += reward

            # Get next state and action for SARSA update
            next_bird_state = bird.get_state(player, knight)
            next_bird_action = bird.sarsa.get_action(next_bird_state)

            # Update Q-table for bird only
            bird.sarsa.update_q_table(bird_state, bird_action, reward, next_bird_state, next_bird_action)

            frame_count += 1

            if not player.alive or not knight.alive:
                break

        bird.end_episode()
        
        if (episode + 1) % 1000 == 0:
            bird.sarsa.save_q_table()
            gc.collect()

        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Bird Epsilon: {bird.sarsa.epsilon:.6f}", flush=True)

    print("Training complete")
    print(f"Final Epsilon: {bird.sarsa.epsilon:.6f}")
    bird.sarsa.save_q_table()

    total_time = (time.time() - start_time) / 60
    print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()
def visualize_bird_knight_training():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    knight = Knight(500, SCREEN_HEIGHT - 72)

    knight.sarsa.q_table = knight.sarsa.load_q_table()
    bird.sarsa.q_table = bird.sarsa.load_q_table()

    num_episodes = 100
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        bird.reset()
        player.reset()
        player.reset_shield()
        knight.reset()
        
        frame_count = 0
        episode_reward = 0

        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            # Knight's turn (using best action, not training)
            knight_state = knight.get_state(player)
            knight_action = knight.sarsa.get_best_action(knight_state)
            knight.act(knight_action, player, tile_map)
            
            # Player's turn
            player.make_decision(knight)
            
            # Bird's turn
            bird_state = bird.get_state(player, knight)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            
            # Update all entities
            player.update(knight, tile_map)
            knight.update(player, tile_map)
            bird.update(player, knight)
            
            # Calculate reward
            reward = bird.get_reward(player, knight)
            episode_reward += reward

            # Get next state and action for SARSA update
            next_bird_state = bird.get_state(player, knight)
            next_bird_action = bird.sarsa.get_action(next_bird_state)

            # Update Q-table for bird only
            bird.sarsa.update_q_table(bird_state, bird_action, reward, next_bird_state, next_bird_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)
            
            # Draw sprites
            screen.blit(player.image, player.rect)
            screen.blit(knight.image, knight.rect)
            screen.blit(bird.image, bird.rect)
            
            # Draw shield
            bird.draw_shield(screen, player)

            # Draw health bars
            # Player health bar
            pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 10, player.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (player.rect.x, player.rect.y - 10, player.rect.width * (player.health / player.max_health), 5))
            
            # Knight health bar
            pygame.draw.rect(screen, RED, (knight.rect.x, knight.rect.y - 10, knight.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (knight.rect.x, knight.rect.y - 10, knight.rect.width * (knight.health / knight.max_health), 5))
            
            # Display episode number, current reward, and actions
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode+1}", True, BLACK)
            reward_text = font.render(f"Reward: {episode_reward:.2f}", True, BLACK)
            bird_action_text = font.render(f"Bird Action: {bird_action}", True, BLACK)
            knight_action_text = font.render(f"Knight Action: {knight_action}", True, BLACK)
            epsilon_text = font.render(f"Epsilon: {bird.sarsa.epsilon:.4f}", True, BLACK)
            screen.blit(episode_text, (10, 10))
            screen.blit(reward_text, (10, 40))
            screen.blit(bird_action_text, (10, 70))
            screen.blit(knight_action_text, (10, 100))
            screen.blit(epsilon_text, (10, 160))

            pygame.display.flip()
            clock.tick(60)  

            frame_count += 1

            if not player.alive or not knight.alive:
                break

        bird.end_episode()
        
        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Bird Epsilon: {bird.sarsa.epsilon:.6f}",  flush=True)

    print("Visualization complete")
    bird.sarsa.save_q_table()

    pygame.quit()


def train_enemy_fast():
    tile_map = TileMap()
    enemy = Enemy(500, SCREEN_HEIGHT - 50)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)

    num_episodes = 500000
    frames_per_episode = 60 * 60  # 60 seconds at 60 FPS
    FULL_RESET_INTERVAL = 50000

    start_time = time.time()

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            gc.collect()
            print(f"Performed full reset at episode {episode}, continuing with epsilon {enemy.sarsa.epsilon:.6f}")

        enemy.reset()
        player.reset()

        frame_count = 0
        episode_reward = 0
        successful_hits = 0

        while frame_count < frames_per_episode:
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_action(enemy_state)
            enemy.act(enemy_action, tile_map)

            player.make_decision(enemy)

            previous_enemy_health = enemy.health

            player.update(enemy, tile_map)
            enemy.update(player, tile_map)

            hit_player, killed_player = enemy.check_arrow_hit(player)
            dx = player.rect.x - enemy.rect.x

            # Calculate reward
            reward = 0


            if abs(dx) < 150:
                reward -= 0.1
            if enemy.health < previous_enemy_health:
                reward -= 40
            if hit_player:
                reward += 60
                successful_hits += 1
            if killed_player:
                reward += 120
            if enemy.health <= 0:
                reward -= 100
            if enemy.just_attacked:
                reward -= 5
                enemy.just_attacked = False
            episode_reward += reward

            # Get next state and action
            next_enemy_state = enemy.get_state(player)
            next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

            # Update Q-table
            enemy.sarsa.update_q_table(enemy_state, enemy_action, reward, next_enemy_state, next_enemy_action)

            frame_count += 1

            if not player.alive or not enemy.alive:
                break

        enemy.end_episode()

        print(
            f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {enemy.sarsa.epsilon:.6f}, Successful Hits: {successful_hits}",
            flush=True)

        if (episode + 1) % 1000 == 0:
            enemy.sarsa.save_q_table()
            gc.collect()

    print("Training complete")
    print(f"Final Epsilon: {enemy.sarsa.epsilon:.6f}")
    enemy.sarsa.save_q_table()

    total_time = (time.time() - start_time) / 60
    print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()
    
    
def visualize_enemy_training():
    tile_map = TileMap()
    enemy = Enemy(500, SCREEN_HEIGHT - 50)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)

    all_sprites = pygame.sprite.Group(player, enemy)

    num_episodes = 100
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        enemy.reset()
        player.reset()

        frame_count = 0
        episode_reward = 0
        successful_hits = 0

        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_action(enemy_state)
            enemy.act(enemy_action, tile_map)

            player.make_decision(enemy)

            previous_enemy_health = enemy.health

            player.update(enemy, tile_map)
            enemy.update(player, tile_map)

            hit_player, killed_player = enemy.check_arrow_hit(player)
            dx = player.rect.x - enemy.rect.x

            # Calculate reward
            reward = 0
            if abs(dx) < 100:
                reward -= 0.05
            if abs(dx) < 150:
                reward -= 0.01
            if enemy.health < previous_enemy_health:
                reward -= 30
            if hit_player:
                reward += 50
                successful_hits += 1
            if killed_player:
                reward += 100
            if enemy.health <= 0:
                reward -= 100
            if enemy.just_attacked:
                reward -= 20
                enemy.just_attacked = False
            episode_reward += reward
            
            # Get next state and action
            next_enemy_state = enemy.get_state(player)
            next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

            # Update Q-table
            enemy.sarsa.update_q_table(enemy_state, enemy_action, reward, next_enemy_state, next_enemy_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)

            # Draw sprites
            all_sprites.draw(screen)

            # Draw enemy arrows
            enemy.draw_arrows(screen)

            # Draw health bars
            pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 10, player.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (
            player.rect.x, player.rect.y - 10, player.rect.width * (player.health / player.max_health), 5))
            pygame.draw.rect(screen, RED, (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width, 5))
            pygame.draw.rect(screen, GREEN,
                             (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width * (enemy.health / enemy.max_health), 5))

            # Display episode number, current reward, epsilon, and successful hits
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode + 1}", True, BLACK)
            reward_text = font.render(f"Reward: {episode_reward:.2f}", True, BLACK)
            epsilon_text = font.render(f"Epsilon: {enemy.sarsa.epsilon:.4f}", True, BLACK)
            hits_text = font.render(f"Successful Hits: {successful_hits}", True, BLACK)

            screen.blit(episode_text, (10, 10))
            screen.blit(reward_text, (10, 40))
            screen.blit(epsilon_text, (10, 70))
            screen.blit(hits_text, (10, 100))

            pygame.display.flip()
            clock.tick(60)

            frame_count += 1

            if not player.alive or not enemy.alive:
                break

        enemy.end_episode()

        print(
            f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {enemy.sarsa.epsilon:.6f}, Successful Hits: {successful_hits}",
            flush=True)

    print("Visualization complete")
    enemy.sarsa.save_q_table()

    pygame.quit()
def train_bird_and_enemy_fast():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    enemy = Enemy(500, SCREEN_HEIGHT - 50)
    
    bird.sarsa.q_table = bird.sarsa.load_q_table()
    enemy.sarsa.q_table = enemy.sarsa.load_q_table()

    num_episodes = 1000000
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS
    FULL_RESET_INTERVAL = 1000

    start_time = time.time()

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            gc.collect()
            print(f"Performed full reset at episode {episode}, continuing with epsilon Bird: {bird.sarsa.epsilon:.6f}")
        
        bird.reset()
        player.reset()
        player.reset_shield()
        enemy.reset()
        enemy.sarsa.epsilon = 0
        frame_count = 0
        bird_episode_reward = 0
        enemy_episode_reward = 0
        
        while frame_count < frames_per_episode:
            # Enemy's turn
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_action(enemy_state)
            enemy.act(enemy_action, tile_map)
            
            # Player's turn
            player.make_decision(enemy)
            
            # Bird's turn
            bird_state = bird.get_state(player, enemy=enemy)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            
            previous_enemy_health = enemy.health
            
            # Update all entities
            player.update(enemy, tile_map)
            enemy.update(player, tile_map)
            bird.update(player, enemy=enemy, knight=None)
            dx = player.rect.x - enemy.rect.x

            hit_player, killed_player = enemy.check_arrow_hit(player)
            # Calculate rewards
            bird_reward = bird.get_reward(player, enemy=enemy)

            
            bird_episode_reward += bird_reward


            # Get next states and actions for SARSA update
            next_bird_state = bird.get_state(player, enemy=enemy)
            next_bird_action = bird.sarsa.get_action(next_bird_state)
            
            next_enemy_state = enemy.get_state(player)
            next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

            # Update Q-tables
            bird.sarsa.update_q_table(bird_state, bird_action, bird_reward, next_bird_state, next_bird_action)
            #enemy.sarsa.update_q_table(enemy_state, enemy_action, enemy_reward, next_enemy_state, next_enemy_action)

            frame_count += 1

            if not player.alive or not enemy.alive:
                break

        bird.end_episode()
        #enemy.end_episode()
        
        if (episode + 1) % 1000 == 0:
            bird.sarsa.save_q_table()

            gc.collect()

        print(f"Episode {episode + 1}: Bird Reward: {bird_episode_reward:.2f} "
              f"Bird Epsilon: {bird.sarsa.epsilon:.6f}", flush=True)

    print("Training complete")
    print(f"Final Epsilon - Bird: {bird.sarsa.epsilon:.6f}")
    bird.sarsa.save_q_table()


    total_time = (time.time() - start_time) / 60
    print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()
    
def visualize_bird_and_enemy_training():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    enemy = Enemy(500, SCREEN_HEIGHT - 50)

    bird.sarsa.q_table = bird.sarsa.load_q_table()
    enemy.sarsa.q_table = enemy.sarsa.load_q_table()

    num_episodes = 100
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        bird.reset()
        player.reset()
        player.reset_shield()
        enemy.reset()
        
        frame_count = 0
        bird_episode_reward = 0
        enemy_episode_reward = 0
        
        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            # Enemy's turn
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_action(enemy_state)
            enemy.act(enemy_action, tile_map)
            
            # Player's turn
            player.make_decision(enemy)
            
            # Bird's turn
            bird_state = bird.get_state(player, enemy=enemy)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            previous_enemy_health = enemy.health
            
            # Update all entities
            player.update(enemy, tile_map)
            enemy.update(player, tile_map)
            bird.update(player, enemy=enemy)
            dx = player.rect.x - enemy.rect.x

            hit_player, killed_player = enemy.check_arrow_hit(player)
            # Calculate rewards
            bird_reward = bird.get_reward(player, enemy=enemy)
            enemy_reward = 0

            if abs(dx) < 100:
                enemy_reward -= 10
            if abs(dx) < 150:
                enemy_reward -= 5
            if enemy.rect.left <= 50 or enemy.rect.right >= SCREEN_WIDTH - 50:
                enemy_reward -= 10
            if enemy.rect.left <= 100 or enemy.rect.right >= SCREEN_WIDTH - 100:
                enemy_reward -= 5
            if enemy.health < previous_enemy_health:
                enemy_reward -= 40
            if hit_player:
                enemy_reward += 40
            if killed_player:
                enemy_reward += 100
            if enemy.health <= 0:
                enemy_reward -= 100
            if enemy.just_attacked:
                enemy_reward -= 10
                enemy.just_attacked = False
            
            bird_episode_reward += bird_reward
            enemy_episode_reward += enemy_reward

            # Get next states and actions for SARSA update
            next_bird_state = bird.get_state(player, enemy=enemy)
            next_bird_action = bird.sarsa.get_action(next_bird_state)
            
            next_enemy_state = enemy.get_state(player)
            next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

            # Update Q-tables
            bird.sarsa.update_q_table(bird_state, bird_action, bird_reward, next_bird_state, next_bird_action)
            enemy.sarsa.update_q_table(enemy_state, enemy_action, enemy_reward, next_enemy_state, next_enemy_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)
            
            # Draw sprites
            screen.blit(player.image, player.rect)
            screen.blit(enemy.image, enemy.rect)
            screen.blit(bird.image, bird.rect)
            
            # Draw enemy arrows
            enemy.draw_arrows(screen)
            
            # Draw shield
            bird.draw_shield(screen, player)

            # Draw health bars
            # Player health bar
            pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 10, player.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (player.rect.x, player.rect.y - 10, player.rect.width * (player.health / player.max_health), 5))
            
            # Enemy health bar
            pygame.draw.rect(screen, RED, (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width * (enemy.health / enemy.max_health), 5))
            
            # Display episode number, current rewards, and actions
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode+1}", True, BLACK)
            bird_reward_text = font.render(f"Bird Reward: {bird_episode_reward:.2f}", True, BLACK)
            enemy_reward_text = font.render(f"Enemy Reward: {enemy_episode_reward:.2f}", True, BLACK)
            bird_action_text = font.render(f"Bird Action: {bird_action}", True, BLACK)
            enemy_action_text = font.render(f"Enemy Action: {enemy_action}", True, BLACK)
            bird_epsilon_text = font.render(f"Bird Epsilon: {bird.sarsa.epsilon:.4f}", True, BLACK)
            enemy_epsilon_text = font.render(f"Enemy Epsilon: {enemy.sarsa.epsilon:.4f}", True, BLACK)
            
            screen.blit(episode_text, (10, 10))
            screen.blit(bird_reward_text, (10, 40))
            screen.blit(enemy_reward_text, (10, 70))
            screen.blit(bird_action_text, (10, 100))
            screen.blit(enemy_action_text, (10, 130))
            screen.blit(bird_epsilon_text, (10, 160))
            screen.blit(enemy_epsilon_text, (10, 190))

            pygame.display.flip()
            clock.tick(60)  

            frame_count += 1

            if not player.alive or not enemy.alive:
                break

        bird.end_episode()
        enemy.end_episode()
        
        print(f"Episode {episode + 1}: Bird Reward: {bird_episode_reward:.2f}, Enemy Reward: {enemy_episode_reward:.2f}, "
              f"Bird Epsilon: {bird.sarsa.epsilon:.6f}, Enemy Epsilon: {enemy.sarsa.epsilon:.6f}", flush=True)

    print("Visualization complete")
    bird.sarsa.save_q_table()
    enemy.sarsa.save_q_table()

    pygame.quit()
    
def visualize_bird_and_enemy_training():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    enemy = Enemy(500, SCREEN_HEIGHT - 50)

    bird.sarsa.q_table = bird.sarsa.load_q_table()
    enemy.sarsa.q_table = enemy.sarsa.load_q_table()

    num_episodes = 100
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        bird.reset()
        player.reset()
        player.reset_shield()
        enemy.reset()
        
        frame_count = 0
        bird_episode_reward = 0
        enemy_episode_reward = 0
        
        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            # Enemy's turn
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_action(enemy_state)
            enemy.act(enemy_action, tile_map)
            
            # Player's turn
            player.make_decision(enemy)
            
            # Bird's turn
            bird_state = bird.get_state(player, enemy=enemy)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            previous_enemy_health = enemy.health
            
            # Update all entities
            player.update(enemy, tile_map)
            enemy.update(player, tile_map)
            bird.update(player, enemy=enemy)
            dx = player.rect.x - enemy.rect.x

            hit_player, killed_player = enemy.check_arrow_hit(player)
            # Calculate rewards
            bird_reward = bird.get_reward(player, enemy=enemy)
            enemy_reward = 0

            if abs(dx) < 100:
                enemy_reward -= 10
            if abs(dx) < 150:
                enemy_reward -= 5
            if enemy.rect.left <= 100 or enemy.rect.right >= SCREEN_WIDTH - 100:
                enemy_reward = -5
            if enemy.rect.left <= 150 or enemy.rect.right >= SCREEN_WIDTH - 150:
                enemy_reward -= 1
            if enemy.health < previous_enemy_health:
                enemy_reward -= 40
            if hit_player:
                enemy_reward += 40
            if killed_player:
                enemy_reward += 100
            if enemy.health <= 0:
                enemy_reward -= 100
            if enemy.just_attacked:
                enemy_reward -= 10
                enemy.just_attacked = False
            
            bird_episode_reward += bird_reward
            enemy_episode_reward += enemy_reward

            # Get next states and actions for SARSA update
            next_bird_state = bird.get_state(player, enemy=enemy)
            next_bird_action = bird.sarsa.get_action(next_bird_state)
            
            next_enemy_state = enemy.get_state(player)
            next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

            # Update Q-tables
            bird.sarsa.update_q_table(bird_state, bird_action, bird_reward, next_bird_state, next_bird_action)
            enemy.sarsa.update_q_table(enemy_state, enemy_action, enemy_reward, next_enemy_state, next_enemy_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)
            
            # Draw sprites
            screen.blit(player.image, player.rect)
            screen.blit(enemy.image, enemy.rect)
            screen.blit(bird.image, bird.rect)
            
            # Draw enemy arrows
            enemy.draw_arrows(screen)
            
            # Draw shield
            bird.draw_shield(screen, player)

            # Draw health bars
            # Player health bar
            pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 10, player.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (player.rect.x, player.rect.y - 10, player.rect.width * (player.health / player.max_health), 5))
            
            # Enemy health bar
            pygame.draw.rect(screen, RED, (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width, 5))
            pygame.draw.rect(screen, GREEN, (enemy.rect.x, enemy.rect.y - 10, enemy.rect.width * (enemy.health / enemy.max_health), 5))
            
            # Display episode number, current rewards, and actions
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode+1}", True, BLACK)
            bird_reward_text = font.render(f"Bird Reward: {bird_episode_reward:.2f}", True, BLACK)
            enemy_reward_text = font.render(f"Enemy Reward: {enemy_episode_reward:.2f}", True, BLACK)
            bird_action_text = font.render(f"Bird Action: {bird_action}", True, BLACK)
            enemy_action_text = font.render(f"Enemy Action: {enemy_action}", True, BLACK)
            bird_epsilon_text = font.render(f"Bird Epsilon: {bird.sarsa.epsilon:.4f}", True, BLACK)
            enemy_epsilon_text = font.render(f"Enemy Epsilon: {enemy.sarsa.epsilon:.4f}", True, BLACK)
            
            screen.blit(episode_text, (10, 10))
            screen.blit(bird_reward_text, (10, 40))
            screen.blit(enemy_reward_text, (10, 70))
            screen.blit(bird_action_text, (10, 100))
            screen.blit(enemy_action_text, (10, 130))
            screen.blit(bird_epsilon_text, (10, 160))
            screen.blit(enemy_epsilon_text, (10, 190))

            pygame.display.flip()
            clock.tick(60)  

            frame_count += 1

            if not player.alive or not enemy.alive:
                break

        bird.end_episode()
        enemy.end_episode()
        
        print(f"Episode {episode + 1}: Bird Reward: {bird_episode_reward:.2f}, Enemy Reward: {enemy_episode_reward:.2f}, "
              f"Bird Epsilon: {bird.sarsa.epsilon:.6f}, Enemy Epsilon: {enemy.sarsa.epsilon:.6f}", flush=True)

    print("Visualization complete")
    bird.sarsa.save_q_table()
    enemy.sarsa.save_q_table()

    pygame.quit()
    
def train_bird_with_knight_and_enemy_fast():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    knight = Knight(500, SCREEN_HEIGHT - 72)
    enemy = Enemy(600, SCREEN_HEIGHT - 50)

    knight.sarsa.epsilon_min = 0
    knight.sarsa.epsilon = 0
    enemy.sarsa.epsilon_min = 0
    enemy.sarsa.epsilon = 0

    knight.sarsa.q_table = knight.sarsa.load_q_table()
    enemy.sarsa.q_table = enemy.sarsa.load_q_table()
    bird.sarsa.q_table = bird.sarsa.load_q_table()

    num_episodes = 1000000
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS
    FULL_RESET_INTERVAL = 1000

    start_time = time.time()

    for episode in range(num_episodes):
        if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
            gc.collect()
            print(f"Performed full reset at episode {episode}, continuing with epsilon {bird.sarsa.epsilon:.6f}")
        
        bird.reset()
        player.reset()
        player.reset_shield()
        knight.reset()
        enemy.reset()
        
        frame_count = 0
        episode_reward = 0
        
        while frame_count < frames_per_episode:
            # Knight's turn (using best action, not training)
            knight_state = knight.get_state(player)
            knight_action = knight.sarsa.get_best_action(knight_state)
            knight.act(knight_action, player, tile_map)
            
            # Enemy's turn (using best action, not training)
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_best_action(enemy_state)
            enemy.act(enemy_action, tile_map)
            
            # Player's turn
            player.make_decision(knight, enemy)
            
            # Bird's turn
            bird_state = bird.get_state(player, knight=knight, enemy=enemy)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            
            # Update all entities
            player.update(knight, enemy, tile_map)
            knight.update(player, tile_map)
            enemy.update(player, tile_map)
            bird.update(player, knight=knight, enemy=enemy)
            
            # Check for arrow hits
            enemy.check_arrow_hit(player)
            
            # Check for player's attack hitting enemy or knight
            if player.attacking and not player.has_hit_enemy:
                if (abs(player.rect.centerx - enemy.rect.centerx) < player.attack_range and
                    abs(player.rect.centery - enemy.rect.centery) < 50):
                    knockback_direction = 1 if player.facing_right else -1
                    enemy.take_damage(10, knockback_direction)
                    player.has_hit_enemy = True
                elif (abs(player.rect.centerx - knight.rect.centerx) < player.attack_range and
                      abs(player.rect.centery - knight.rect.centery) < 50):
                    knockback_direction = 1 if player.facing_right else -1
                    knight.take_damage(10, knockback_direction)
                    player.has_hit_enemy = True

            # Check for knight's attack hitting player
            knight.check_melee_hit(player)
            
            # Calculate reward
            reward = bird.get_reward(player, knight=knight, enemy=enemy)
            episode_reward += reward

            # Get next state and action for SARSA update
            next_bird_state = bird.get_state(player, knight=knight, enemy=enemy)
            next_bird_action = bird.sarsa.get_action(next_bird_state)

            # Update Q-table for bird only
            bird.sarsa.update_q_table(bird_state, bird_action, reward, next_bird_state, next_bird_action)

            frame_count += 1

            if not player.alive or (not knight.alive and not enemy.alive):
                break

        bird.end_episode()
        
        if (episode + 1) % 1000 == 0:
            bird.sarsa.save_q_table()
            gc.collect()

        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Bird Epsilon: {bird.sarsa.epsilon:.6f}", flush=True)

    print("Training complete")
    print(f"Final Epsilon: {bird.sarsa.epsilon:.6f}")
    bird.sarsa.save_q_table()

    total_time = (time.time() - start_time) / 60
    print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()

def visualize_bird_knight_and_enemy_training():
    tile_map = TileMap()
    bird = Bird(400, SCREEN_HEIGHT - 100)
    player = AIPlayer(250, SCREEN_HEIGHT - 50)
    knight = Knight(500, SCREEN_HEIGHT - 72)
    enemy = Enemy(600, SCREEN_HEIGHT - 50)

    knight.sarsa.q_table = knight.sarsa.load_q_table()
    enemy.sarsa.q_table = enemy.sarsa.load_q_table()
    bird.sarsa.q_table = bird.sarsa.load_q_table()

    num_episodes = 100
    frames_per_episode = 30 * 60  # 30 seconds at 60 FPS

    clock = pygame.time.Clock()

    for episode in range(num_episodes):
        bird.reset()
        player.reset()
        player.reset_shield()
        knight.reset()
        enemy.reset()
        
        frame_count = 0
        episode_reward = 0

        while frame_count < frames_per_episode:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()

            # Knight's turn (using best action, not training)
            knight_state = knight.get_state(player)
            knight_action = knight.sarsa.get_best_action(knight_state)
            knight.act(knight_action, player, tile_map)
            
            # Enemy's turn (using best action, not training)
            enemy_state = enemy.get_state(player)
            enemy_action = enemy.sarsa.get_best_action(enemy_state)
            enemy.act(enemy_action, tile_map)
            
            # Player's turn
            if knight.alive:
                player.make_decision(knight)
            else:
                player.make_decision(enemy)
            # Bird's turn
            bird_state = bird.get_state(player, knight=knight, enemy=enemy)
            bird_action = bird.sarsa.get_action(bird_state)
            bird.perform_action(bird_action, player)
            
            # Update all entities
            if enemy.alive:
                player.update(enemy, tile_map)
            else:
                player.update(knight, tile_map)
            knight.update(player, tile_map)
            enemy.update(player, tile_map)
            bird.update(player, knight=knight, enemy=enemy)
            
            # Check for arrow hits
            enemy.check_arrow_hit(player)
            
            # Check for player's attack hitting enemy or knight
            if player.attacking and not player.has_hit_enemy:
                if (abs(player.rect.centerx - enemy.rect.centerx) < player.attack_range and
                    abs(player.rect.centery - enemy.rect.centery) < 50):
                    knockback_direction = 1 if player.facing_right else -1
                    enemy.take_damage(10, knockback_direction)
                    player.has_hit_enemy = True
                elif (abs(player.rect.centerx - knight.rect.centerx) < player.attack_range and
                      abs(player.rect.centery - knight.rect.centery) < 50):
                    knockback_direction = 1 if player.facing_right else -1
                    knight.take_damage(10, knockback_direction)
                    player.has_hit_enemy = True

            # Check for knight's attack hitting player
            knight.check_melee_hit(player)
            
            # Calculate reward
            reward = bird.get_reward(player, knight=knight, enemy=enemy)
            episode_reward += reward

            # Get next state and action for SARSA update
            next_bird_state = bird.get_state(player, knight=knight, enemy=enemy)
            next_bird_action = bird.sarsa.get_action(next_bird_state)

            # Update Q-table for bird only
            bird.sarsa.update_q_table(bird_state, bird_action, reward, next_bird_state, next_bird_action)

            # Drawing
            screen.fill(WHITE)
            tile_map.draw(screen)
            
            # Draw sprites
            screen.blit(player.image, player.rect)
            screen.blit(knight.image, knight.rect)
            screen.blit(enemy.image, enemy.rect)
            screen.blit(bird.image, bird.rect)
            
            # Draw shield
            bird.draw_shield(screen, player)

            # Draw enemy arrows
            enemy.draw_arrows(screen)

            # Draw health bars
            for character in [player, knight, enemy]:
                pygame.draw.rect(screen, RED, (character.rect.x, character.rect.y - 10, character.rect.width, 5))
                pygame.draw.rect(screen, GREEN, (character.rect.x, character.rect.y - 10, character.rect.width * (character.health / character.max_health), 5))
            
            # Display episode number, current reward, and actions
            font = pygame.font.Font(None, 24)
            episode_text = font.render(f"Episode: {episode+1}", True, BLACK)
            reward_text = font.render(f"Reward: {episode_reward:.2f}", True, BLACK)
            bird_action_text = font.render(f"Bird Action: {bird_action}", True, BLACK)
            knight_action_text = font.render(f"Knight Action: {knight_action}", True, BLACK)
            enemy_action_text = font.render(f"Enemy Action: {enemy_action}", True, BLACK)
            epsilon_text = font.render(f"Epsilon: {bird.sarsa.epsilon:.4f}", True, BLACK)
            screen.blit(episode_text, (10, 10))
            screen.blit(reward_text, (10, 40))
            screen.blit(bird_action_text, (10, 70))
            screen.blit(knight_action_text, (10, 100))
            screen.blit(enemy_action_text, (10, 130))
            screen.blit(epsilon_text, (10, 160))

            pygame.display.flip()
            clock.tick(60)  

            frame_count += 1

            if not player.alive or (not knight.alive and not enemy.alive):
                break

        bird.end_episode()
        
        print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Bird Epsilon: {bird.sarsa.epsilon:.6f}",  flush=True)

    print("Visualization complete")
    bird.sarsa.save_q_table()

    pygame.quit()
    
def main():
    clock = pygame.time.Clock()
    tile_map = TileMap()
    player = Player(250, SCREEN_HEIGHT - 100)
    enemy = Enemy(500, SCREEN_HEIGHT - 100)
    knight = Knight(700, SCREEN_HEIGHT - 100)
    bird = Bird(400, SCREEN_HEIGHT - 150)
    
    all_sprites = pygame.sprite.Group(player, enemy, knight, bird)
    bird.sarsa.epsilon = 0
    enemy.sarsa.epsilon = 0
    knight.sarsa.epsilon = 0
    running = True
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_UP:
                    player.jump()
                if event.key == pygame.K_SPACE:
                    player.attack()

        keys = pygame.key.get_pressed()
        if keys[pygame.K_LEFT]:
            player.move(-player.speed, tile_map)
        if keys[pygame.K_RIGHT]:
            player.move(player.speed, tile_map)

        # Update all entities
        player.update(tile_map)
        enemy.update(player, tile_map)
        knight.update(player, tile_map)
        bird.update(player, enemy, knight)

        # Check for collisions between player and arrows
        for arrow in enemy.arrow_group:
            if pygame.sprite.collide_rect(arrow, player):
                player.take_damage(5, 1 if arrow.direction > 0 else -1)
                arrow.kill()

        # Check for player's attack hitting enemy or knight
        if player.attacking and not player.has_hit_enemy:
            if (abs(player.rect.centerx - enemy.rect.centerx) < player.attack_range and
                abs(player.rect.centery - enemy.rect.centery) < 50):
                knockback_direction = 1 if player.facing_right else -1
                enemy.take_damage(10, knockback_direction)
                player.has_hit_enemy = True
            elif (abs(player.rect.centerx - knight.rect.centerx) < player.attack_range and
                  abs(player.rect.centery - knight.rect.centery) < 50):
                knockback_direction = 1 if player.facing_right else -1
                knight.take_damage(10, knockback_direction)
                player.has_hit_enemy = True

        # Check for knight's attack hitting player
        if knight.attacking and not knight.attack_landed:
            if (abs(knight.rect.centerx - player.rect.centerx) < knight.attack_range and
                abs(knight.rect.centery - player.rect.centery) < 50):
                knockback_direction = 1 if knight.direction > 0 else -1
                player.take_damage(10, knockback_direction)
                knight.attack_landed = True

        # Drawing
        screen.fill(WHITE)
        tile_map.draw(screen)
        all_sprites.draw(screen)
        enemy.draw_arrows(screen)
        bird.draw_shield(screen, player)

        # Draw health bars
        pygame.draw.rect(screen, RED, (player.rect.x, player.rect.y - 20, player.rect.width, 5))
        pygame.draw.rect(screen, GREEN, (player.rect.x, player.rect.y - 20, player.rect.width * player.health / player.max_health, 5))

        pygame.draw.rect(screen, RED, (enemy.rect.x, enemy.rect.y - 20, enemy.rect.width, 5))
        pygame.draw.rect(screen, GREEN, (enemy.rect.x, enemy.rect.y - 20, enemy.rect.width * enemy.health / enemy.max_health, 5))

        pygame.draw.rect(screen, RED, (knight.rect.x, knight.rect.y - 20, knight.rect.width, 5))
        pygame.draw.rect(screen, GREEN, (knight.rect.x, knight.rect.y - 20, knight.rect.width * knight.health / knight.max_health, 5))

        # Display game information
        font = pygame.font.Font(None, 24)
        player_health_text = font.render(f"Player Health: {player.health}", True, BLACK)
        enemy_health_text = font.render(f"Enemy Health: {enemy.health}", True, BLACK)
        knight_health_text = font.render(f"Knight Health: {knight.health}", True, BLACK)
        
        screen.blit(player_health_text, (10, 10))
        screen.blit(enemy_health_text, (10, 40))
        screen.blit(knight_health_text, (10, 70))

        pygame.display.flip()
        clock.tick(FPS)

    pygame.quit()
    
def show_map():
    """
    Display only the tile map without any characters.
    Press ESC or close button to exit.
    """
    pygame.init()
    clock = pygame.time.Clock()
    tile_map = TileMap()
    player = Player(250, SCREEN_HEIGHT - 100)
    all_sprites = pygame.sprite.Group(player)
    running = True
    running = True
    while running:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                running = False
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_UP:
                    player.jump()
                if event.key == pygame.K_SPACE:
                    player.attack()

        keys = pygame.key.get_pressed()
        if keys[pygame.K_LEFT]:
            player.move(-player.speed, tile_map)
        if keys[pygame.K_RIGHT]:
            player.move(player.speed, tile_map)

        # Update all entities
        player.update(tile_map)

        # Clear screen

        screen.fill(WHITE)
        all_sprites.draw(screen)
        # Draw tile map
        tile_map.draw(screen)
        
            
       
        pygame.display.flip()
        clock.tick(60)

    pygame.quit()

    
def test_knight_performance():
    import matplotlib.pyplot as plt
    import numpy as np

    q_table_numbers = list(range(0, 50001, 100))
    x_data = []
    y_data = []

    frames_per_episode = 1000  # Number of frames per episode

    for q_table_number in q_table_numbers:
        total_rewards = []
        print(f"Testing q_table_number: {q_table_number}")

        # Initialize tile_map, knight, player
        tile_map = TileMap()
        knight = Knight(500, SCREEN_HEIGHT - 72)
        player = AIPlayer(250, SCREEN_HEIGHT - 50)

        # Initialize knight's SARSA
        if q_table_number == 0:
            knight.sarsa = SARSA(character_type="knight")
            knight.sarsa.q_table = {}  # Empty q_table
            knight.sarsa.epsilon = 1  # Fully random actions
        else:
            knight.sarsa = SARSA(character_type="knight")
            # Load q_table from file
            q_table_file = f'knight_q_tables/q_table_episode_{q_table_number}.json'
            if not os.path.exists(q_table_file):
                print(f"Q-table file {q_table_file} not found. Skipping.")
                continue
            with open(q_table_file, 'r') as f:
                knight.sarsa.q_table = json.load(f)
            knight.sarsa.epsilon = 0  # Greedy policy

        # Run 100 episodes
        for episode in range(100):
            knight.reset()
            player.reset()
            total_reward = 0
            frame_count = 0

            while frame_count < frames_per_episode:
                current_state = knight.get_state(player)
                action = knight.sarsa.get_action(current_state)

                previous_health = knight.health
                knight.act(action, player, tile_map)
                player.update(knight, tile_map)
                knight.update(player, tile_map)
                dx = player.rect.x - knight.rect.x

                # Calculate reward
                reward = 0
                if knight.just_attacked:
                    reward -= 10
                    knight.just_attacked = False

                if abs(dx) < 100:
                    reward += 0.01
                if knight.hit_player:
                    reward += 40
                    if knight.killed_player:
                        reward += 100
                if knight.health < previous_health:
                    reward -= 40
                if knight.blocking and knight.shield_used and knight.is_facing_player():
                    reward += 30
                if knight.health == 0 and not knight.death_penalty_applied:
                    reward -= 100
                    knight.death_penalty_applied = True

                total_reward += reward

                frame_count += 1

                if not player.alive or not knight.alive:
                    break

            total_rewards.append(total_reward)

        # Compute average rewards every 10 episodes
        avg_rewards_10 = []
        for i in range(0, 100, 10):
            avg = sum(total_rewards[i:i+10]) / 10
            avg_rewards_10.append(avg)

        # Append data
        x_data.extend([q_table_number]*10)
        y_data.extend(avg_rewards_10)

        print(f"Average reward for q_table {q_table_number}: {sum(total_rewards)/len(total_rewards):.2f}")

    # Now plot the results
    plt.figure(figsize=(10, 6))
    plt.scatter(x_data, y_data, alpha=0.5)
    plt.xlabel('Training Episodes')
    plt.ylabel('Average Reward over 10 Episodes')
    plt.title('Knight Performance over Training')
    plt.grid(True)

    # Add logarithmic trend line
    x_array = np.array(x_data)
    y_array = np.array(y_data)

    # Exclude q_table_number == 0 from trendline calculation
    mask = x_array > 0
    x_array = x_array[mask]
    y_array = y_array[mask]

    coefficients = np.polyfit(np.log(x_array), y_array, 1)
    trendline = coefficients[0] * np.log(x_array) + coefficients[1]
    plt.plot(x_array, trendline, label='Logarithmic Trendline', color='red')
    plt.legend()

    plt.show()

def test():
    import pandas as pd
    from openpyxl import load_workbook
    import os
    for i in range(2):
        excel_file = 'enemy_training_rewards.xlsx'
        
        # Load existing data or create new DataFrame
        if os.path.exists(excel_file):
            try:
                df = pd.read_excel(excel_file)
                # Find the first empty column
                next_col = len(df.columns)
                column_name = f'Run_{next_col + 1}'
            except Exception as e:
                print(f"Error reading existing Excel file: {e}")
                df = pd.DataFrame()
                column_name = 'Run_1'
        else:
            df = pd.DataFrame()
            column_name = 'Run_1'

        rewards_list = []  # Store rewards for this run

        tile_map = TileMap()
        enemy = Enemy(500, SCREEN_HEIGHT - 50)
        # enemy.sarsa.q_table = enemy.sarsa.load_q_table()  # Add this line
        # enemy.sarsa.epsilon = 0  # Set epsilon to 0 for pure exploitation
        player = AIPlayer(250, SCREEN_HEIGHT - 50)

        num_episodes = 100
        frames_per_episode = 60 * 60  # 30 seconds at 60 FPS
        FULL_RESET_INTERVAL = 50000

        start_time = time.time()
        
        for episode in range(num_episodes):
            if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
                gc.collect()
                print(f"Performed full reset at episode {episode}, continuing with epsilon {enemy.sarsa.epsilon:.6f}")

            enemy.reset()
            player.reset()

            frame_count = 0
            episode_reward = 0
            successful_hits = 0

            while frame_count < frames_per_episode:
                enemy_state = enemy.get_state(player)
                enemy_action = enemy.sarsa.get_action(enemy_state)
                enemy.act(enemy_action, tile_map)

                player.make_decision(enemy)

                previous_enemy_health = enemy.health

                player.update(enemy, tile_map)
                enemy.update(player, tile_map)

                hit_player, killed_player = enemy.check_arrow_hit(player)
                dx = player.rect.x - enemy.rect.x

                # Calculate reward
                reward = 0

                if abs(dx) < 150:
                    reward -= 0.1
                if enemy.health < previous_enemy_health:
                    reward -= 40
                if hit_player:
                    reward += 60
                    successful_hits += 1
                if killed_player:
                    reward += 120
                if enemy.health <= 0:
                    reward -= 100
                if enemy.just_attacked:
                    reward -= 5
                    enemy.just_attacked = False
                episode_reward += reward

                episode_reward += reward

                # Get next state and action
                #next_enemy_state = enemy.get_state(player)
                #next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

                # Update Q-table
                    # Update Q-table
                #enemy.sarsa.update_q_table(enemy_state, enemy_action, reward, next_enemy_state, next_enemy_action)

                frame_count += 1

                if not player.alive or not enemy.alive:
                    break

            enemy.end_episode()
            
            # Add episode reward to rewards list
            rewards_list.append(episode_reward)

            print(
                f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {enemy.sarsa.epsilon:.6f}, Successful Hits: {successful_hits}",
                flush=True)

            #if (episode + 1) % 1000 == 0:
            #    enemy.sarsa.save_q_table()
            #    gc.collect()

        print("Training complete")
        print(f"Final Epsilon: {enemy.sarsa.epsilon:.6f}")
        #enemy.sarsa.save_q_table()
        print(f"episode {enemy.sarsa.episode_count}")

        # Save rewards to Excel
        new_df = pd.DataFrame({column_name: rewards_list})
        
        if df.empty:
            df = new_df
        else:
            # Extend shorter DataFrame with NaN values to match the longer one
            max_length = max(len(df), len(new_df))
            df = df.reindex(range(max_length))
            new_df = new_df.reindex(range(max_length))
            df = pd.concat([df, new_df], axis=1)

        # Save to Excel
        df.to_excel(excel_file, index=False)
        print(f"Rewards saved to {excel_file}")

        total_time = (time.time() - start_time) / 60
        print(f"\nTotal training time: {total_time:.2f} minutes")

    gc.collect()
    
    import pandas as pd
    from openpyxl import load_workbook
    import os
    for i in range(2):
        excel_file = 'enemy_training_rewards.xlsx'
        
        # Load existing data or create new DataFrame
        if os.path.exists(excel_file):
            try:
                df = pd.read_excel(excel_file)
                # Find the first empty column
                next_col = len(df.columns)
                column_name = f'Run_{next_col + 1}'
            except Exception as e:
                print(f"Error reading existing Excel file: {e}")
                df = pd.DataFrame()
                column_name = 'Run_1'
        else:
            df = pd.DataFrame()
            column_name = 'Run_1'

        rewards_list = []  # Store rewards for this run

        tile_map = TileMap()
        enemy = Enemy(500, SCREEN_HEIGHT - 50)
        # enemy.sarsa.q_table = enemy.sarsa.load_q_table()  # Add this line
        # enemy.sarsa.epsilon = 0  # Set epsilon to 0 for pure exploitation
        player = AIPlayer(250, SCREEN_HEIGHT - 50)

        num_episodes = 100
        frames_per_episode = 60 * 60  # 30 seconds at 60 FPS
        FULL_RESET_INTERVAL = 50000

        start_time = time.time()
        
        for episode in range(num_episodes):
            if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
                gc.collect()
                print(f"Performed full reset at episode {episode}, continuing with epsilon {enemy.sarsa.epsilon:.6f}")

            enemy.reset()
            player.reset()

            frame_count = 0
            episode_reward = 0
            successful_hits = 0

            while frame_count < frames_per_episode:
                enemy_state = enemy.get_state(player)
                enemy_action = enemy.sarsa.get_action(enemy_state)
                enemy.act(enemy_action, tile_map)

                player.make_decision(enemy)

                previous_enemy_health = enemy.health

                player.update(enemy, tile_map)
                enemy.update(player, tile_map)

                hit_player, killed_player = enemy.check_arrow_hit(player)
                dx = player.rect.x - enemy.rect.x

                # Calculate reward
                reward = 0

                if abs(dx) < 150:
                    reward -= 0.1
                if enemy.health < previous_enemy_health:
                    reward -= 40
                if hit_player:
                    reward += 60
                    successful_hits += 1
                if killed_player:
                    reward += 120
                if enemy.health <= 0:
                    reward -= 100
                if enemy.just_attacked:
                    reward -= 5
                    enemy.just_attacked = False
                episode_reward += reward

                episode_reward += reward

                # Get next state and action
                #next_enemy_state = enemy.get_state(player)
                #next_enemy_action = enemy.sarsa.get_action(next_enemy_state)

                # Update Q-table
                    # Update Q-table
                #enemy.sarsa.update_q_table(enemy_state, enemy_action, reward, next_enemy_state, next_enemy_action)

                frame_count += 1

                if not player.alive or not enemy.alive:
                    break

            enemy.end_episode()
            
            # Add episode reward to rewards list
            rewards_list.append(episode_reward)

            print(
                f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Epsilon: {enemy.sarsa.epsilon:.6f}, Successful Hits: {successful_hits}",
                flush=True)

            #if (episode + 1) % 1000 == 0:
            #    enemy.sarsa.save_q_table()
            #    gc.collect()

        print("Training complete")
        print(f"Final Epsilon: {enemy.sarsa.epsilon:.6f}")
        #enemy.sarsa.save_q_table()
        print(f"episode {enemy.sarsa.episode_count}")

        # Save rewards to Excel
        new_df = pd.DataFrame({column_name: rewards_list})
        
        if df.empty:
            df = new_df
        else:
            # Extend shorter DataFrame with NaN values to match the longer one
            max_length = max(len(df), len(new_df))
            df = df.reindex(range(max_length))
            new_df = new_df.reindex(range(max_length))
            df = pd.concat([df, new_df], axis=1)

        # Save to Excel
        df.to_excel(excel_file, index=False)
        print(f"Rewards saved to {excel_file}")

        total_time = (time.time() - start_time) / 60
        print(f"\nTotal training time: {total_time:.2f} minutes")
    gc.collect()



    
def test_bird():
    import pandas as pd
    from openpyxl import load_workbook
    import os
    for j in range(1):
        for i in range (2):
            excel_file = 'bird_training_rewards.xlsx'
            
            # Load existing data or create new DataFrame
            if os.path.exists(excel_file):
                try:
                    df = pd.read_excel(excel_file)
                    # Find the first empty column
                    next_col = len(df.columns)
                    column_name = f'Run_{next_col + 1}'
                except Exception as e:
                    print(f"Error reading existing Excel file: {e}")
                    df = pd.DataFrame()
                    column_name = 'Run_1'
            else:
                df = pd.DataFrame()
                column_name = 'Run_1'

            rewards_list = []  # Store rewards for this run

            tile_map = TileMap()
            bird = Bird(400, SCREEN_HEIGHT - 100)
            player = AIPlayer(250, SCREEN_HEIGHT - 50)
            knight = Knight(500, SCREEN_HEIGHT - 72)
            enemy = Enemy(600, SCREEN_HEIGHT - 50)

            knight.sarsa.epsilon_min = 0
            knight.sarsa.epsilon = 0
            enemy.sarsa.epsilon_min = 0
            enemy.sarsa.epsilon = 0
            
            knight.sarsa.q_table = knight.sarsa.load_q_table()
            with open(f"C:\\Users\\mikol\\Desktop\\Reinforcement Learning\\bird_q_tables\\q_table_episode_{1000}.json", 'r') as f:
                bird.sarsa.q_table = json.load(f)
            enemy.sarsa.q_table = enemy.sarsa.load_q_table()
            # = bird.sarsa.load_q_table()
            
            num_episodes = 100
            frames_per_episode = 30 * 60
            FULL_RESET_INTERVAL = 1000

            start_time = time.time()
        
            for episode in range(num_episodes):
                print(f"Episode {j+episode}")
                if episode % FULL_RESET_INTERVAL == 0 and episode > 0:
                    gc.collect()
                    print(f"Performed full reset at episode {episode}, continuing with epsilon {bird.sarsa.epsilon:.6f}")
                
                bird.reset()
                player.reset()
                player.reset_shield()
                knight.reset()
                enemy.reset()
                
                frame_count = 0
                episode_reward = 0
                
                while frame_count < frames_per_episode:
                    # Knight's turn (using best action, not training)
                    knight_state = knight.get_state(player)
                    knight_action = knight.sarsa.get_best_action(knight_state)
                    knight.act(knight_action, player, tile_map)
                    
                    # Enemy's turn (using best action, not training)
                    enemy_state = enemy.get_state(player)
                    enemy_action = enemy.sarsa.get_best_action(enemy_state)
                    enemy.act(enemy_action, tile_map)
                    
                    # Player's turn
                    if knight.health > 0 :
                        player.make_decision(knight)
                    else:
                        player.make_decision(enemy)
                    
                    # Bird's turn
                    bird_state = bird.get_state(player, knight=knight, enemy=enemy)
                    bird_action = bird.sarsa.get_action(bird_state)
                    bird.perform_action(bird_action, player)
                    
                    # Update all entities
                    if knight.health > 0:
                        
                        player.update(knight, tile_map)
                    else:
                        player.update(enemy, tile_map)
                    knight.update(player, tile_map)
                    enemy.update(player, tile_map)
                    bird.update(player, knight=knight, enemy=enemy)
                    
                    # Check for arrow hits
                    enemy.check_arrow_hit(player)
                    
                    # Check for player's attack hitting enemy or knight
                    if player.attacking and not player.has_hit_enemy:
                        if (abs(player.rect.centerx - enemy.rect.centerx) < player.attack_range and
                            abs(player.rect.centery - enemy.rect.centery) < 50):
                            knockback_direction = 1 if player.facing_right else -1
                            enemy.take_damage(10, knockback_direction)
                            player.has_hit_enemy = True
                        elif (abs(player.rect.centerx - knight.rect.centerx) < player.attack_range and
                            abs(player.rect.centery - knight.rect.centery) < 50):
                            knockback_direction = 1 if player.facing_right else -1
                            knight.take_damage(10, knockback_direction)
                            player.has_hit_enemy = True

                    # Check for knight's attack hitting player
                    knight.check_melee_hit(player)
                    
                    # Calculate reward
                    reward = bird.get_reward(player, knight=knight, enemy=enemy)
                    episode_reward += reward

                    # Get next state and action for SARSA update
                    next_bird_state = bird.get_state(player, knight=knight, enemy=enemy)
                    next_bird_action = bird.sarsa.get_action(next_bird_state)

                    # Update Q-table for bird only
                    #bird.sarsa.update_q_table(bird_state, bird_action, reward, next_bird_state, next_bird_action)

                    frame_count += 1

                    if not player.alive or (not knight.alive and not enemy.alive):
                        break

                bird.end_episode()
                
                # Add episode reward to rewards list
                rewards_list.append(episode_reward)
                
                #if (episode + 1) % 1000 == 0:
                #    bird.sarsa.save_q_table()
                #    gc.collect()

                print(f"Episode {episode + 1}: Reward: {episode_reward:.2f}, Bird Epsilon: {bird.sarsa.epsilon:.6f}", flush=True)

            print("Training complete")
            print(f"Final Epsilon: {bird.sarsa.epsilon:.6f}")
            #bird.sarsa.save_q_table()

            # Save rewards to Excel
            new_df = pd.DataFrame({column_name: rewards_list})
            
            if df.empty:
                df = new_df
            else:
                # Extend shorter DataFrame with NaN values to match the longer one
                max_length = max(len(df), len(new_df))
                df = df.reindex(range(max_length))
                new_df = new_df.reindex(range(max_length))
                df = pd.concat([df, new_df], axis=1)

            # Save to Excel
            df.to_excel(excel_file, index=False)
            print(f"Rewards saved to {excel_file}")

            total_time = (time.time() - start_time) / 60
            print(f"\nTotal training time: {total_time:.2f} minutes")

            gc.collect()
if __name__ == "__main__":
    #test()
    #test_bird()
    #test_knight_performance()
    # Uncomment the function you want to run
    #show_map()
    #train_enemy_fast()
    #visualize_enemy_training()
    
    #train_bird_with_knight_fast()
    #visualize_bird_knight_training()
    
    #train_knight_fast()
    #visualize_training()
    
    #train_bird_and_enemy_fast()
    #visualize_bird_and_enemy_training()
    
    #train_bird_with_knight_and_enemy_fast()
    visualize_bird_knight_and_enemy_training()


    #train_knight_with_simple_player()
    #train_enemy_with_simple_player()
    #train_bird_with_simple_player()
    
    #visualize_knight_simple_player_training()
    #visualize_enemy_simple_player_training()
    #visualize_bird_simple_player_training()
    #main()
    